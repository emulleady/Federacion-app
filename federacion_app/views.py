from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import date, datetime

from django.db.models import Q
from django.core.mail import send_mail
from django.conf import settings
from .models import Persona, SolicitudPase, Vinculo, DocumentoSolicitud, Club, Categoria, DocumentoPersona
from .forms import SolicitudPaseForm, ResolucionSolicitudForm


def es_delegado(usuario):
    return usuario.is_authenticated and usuario.rol == "delegado"


def es_federacion(usuario):
    return usuario.is_authenticated and usuario.rol == "federacion"


def es_consejo_disciplina(usuario):
    return usuario.is_authenticated and usuario.rol == "consejo_disciplina"


def _notificar_pedido_jugador(solicitud):
    """
    Le avisa por email a los delegados del club de origen que otro club
    pidió a uno de sus jugadores. Si el envío falla (mal configurado el
    email, sin internet, etc.) no rompe el flujo: solo lo ignora.
    """
    from .models import Usuario
    destinatarios = list(
        Usuario.objects.filter(rol="delegado", club=solicitud.club_origen)
        .exclude(email="").values_list("email", flat=True)
    )
    if not destinatarios:
        return
    try:
        send_mail(
            subject=f"[Federación] {solicitud.club_destino} pidió a {solicitud.persona}",
            message=(
                f"Hola,\n\n"
                f"El club {solicitud.club_destino} solicitó el pase de {solicitud.persona} "
                f"(documento {solicitud.persona.documento}), que figura en tu club, {solicitud.club_origen}.\n\n"
                f"Entrá al sistema para revisar y liberar o rechazar la solicitud:\n"
                f"http://127.0.0.1:8000/solicitudes/liberar/\n\n"
                f"— Federación Fueguina de Futsal"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=destinatarios,
            fail_silently=True,
        )
    except Exception:
        pass


def bienvenida(request):
    """Pantalla pública de entrada, con logo y foto de fondo de la federación."""
    if request.user.is_authenticated:
        return redirect("home")
    return render(request, "federacion_app/bienvenida.html")


@login_required
def home(request):
    """Después del login, manda a cada usuario a su pantalla principal."""
    if request.user.rol == "delegado":
        return redirect("mis_solicitudes")
    elif request.user.rol == "federacion":
        return redirect("panel_solicitudes")
    elif request.user.rol == "consejo_disciplina":
        return redirect("panel_disciplina")
    return redirect("buscar_persona")


# ---------------------------------------------------------------------
# PANTALLA 1: el delegado carga una nueva solicitud
# ---------------------------------------------------------------------

@login_required
@user_passes_test(es_delegado)
def nueva_solicitud(request):
    if request.method == "POST":
        form = SolicitudPaseForm(request.POST, request.FILES, club=request.user.club)
        if form.is_valid():
            documento = form.cleaned_data["documento"]

            # Busca la persona o la crea si es un alta nueva
            persona, creada = Persona.objects.get_or_create(
                documento=documento,
                defaults={
                    "tipo": "jugador",
                    "nombre": form.cleaned_data.get("nombre", ""),
                    "apellido": form.cleaned_data.get("apellido", ""),
                    "fecha_nacimiento": form.cleaned_data.get("fecha_nacimiento") or date.today(),
                    "foto": form.cleaned_data.get("foto"),
                    "numero_carnet": form.cleaned_data.get("numero_carnet", ""),
                    "requiere_carnet": form.cleaned_data.get("requiere_carnet", False),
                },
            )

            club_origen = persona.club_actual if not creada else None

            # Si es un pase real entre dos clubes distintos, primero tiene
            # que liberarlo el club de origen. Si es alta nueva o cambio de
            # categoría dentro del mismo club, va directo a la federación.
            if club_origen and club_origen != request.user.club:
                estado_inicial = "pendiente_liberacion"
            else:
                estado_inicial = "pendiente"

            solicitud = SolicitudPase.objects.create(
                tipo=form.cleaned_data["tipo"],
                tipo_pase=form.cleaned_data.get("tipo_pase", ""),
                persona=persona,
                club_origen=club_origen,
                club_destino=request.user.club,
                categoria_destino=form.cleaned_data.get("categoria_destino"),
                creado_por=request.user,
                estado=estado_inicial,
            )

            archivo = form.cleaned_data.get("archivo_documentacion")
            if archivo:
                DocumentoSolicitud.objects.create(solicitud=solicitud, archivo=archivo)

            formulario_10 = form.cleaned_data.get("formulario_10_firmado")
            if formulario_10:
                DocumentoSolicitud.objects.create(
                    solicitud=solicitud, archivo=formulario_10, descripcion="Formulario 10 firmado"
                )

            if estado_inicial == "pendiente_liberacion":
                _notificar_pedido_jugador(solicitud)

            messages.success(request, "Solicitud enviada. La federación la va a revisar.")
            return redirect("mis_solicitudes")
    else:
        initial = {}
        if request.GET.get("documento"):
            initial["documento"] = request.GET["documento"]
        if request.GET.get("tipo"):
            initial["tipo"] = request.GET["tipo"]
        if request.GET.get("nombre"):
            initial["nombre"] = request.GET["nombre"]
        if request.GET.get("apellido"):
            initial["apellido"] = request.GET["apellido"]
        if request.GET.get("fecha_nacimiento"):
            initial["fecha_nacimiento"] = request.GET["fecha_nacimiento"]
        if request.GET.get("numero_carnet"):
            initial["numero_carnet"] = request.GET["numero_carnet"]
        form = SolicitudPaseForm(club=request.user.club, initial=initial)

    return render(request, "federacion_app/nueva_solicitud.html", {"form": form})


@login_required
@user_passes_test(es_delegado)
def mis_solicitudes(request):
    """Bandeja: solo las solicitudes que todavía están en trámite."""
    solicitudes = SolicitudPase.objects.filter(
        club_destino=request.user.club
    ).exclude(estado__in=["aprobado", "rechazado"]).select_related("persona", "club_origen", "club_destino")
    return render(request, "federacion_app/mis_solicitudes.html", {"solicitudes": solicitudes})


@login_required
@user_passes_test(es_delegado)
def historial_mis_solicitudes(request):
    """Historial: solicitudes ya resueltas (aprobadas o rechazadas)."""
    solicitudes = SolicitudPase.objects.filter(
        club_destino=request.user.club, estado__in=["aprobado", "rechazado"]
    ).select_related("persona", "club_origen", "club_destino")
    return render(request, "federacion_app/historial_solicitudes.html", {"solicitudes": solicitudes})


@login_required
@user_passes_test(es_delegado)
def solicitudes_a_liberar(request):
    """
    Bandeja del delegado del club de ORIGEN: acá ve los pedidos que
    otros clubes hicieron por jugadores suyos, y puede liberarlos o
    rechazarlos antes de que pasen a la federación.
    """
    solicitudes = SolicitudPase.objects.filter(
        club_origen=request.user.club, estado="pendiente_liberacion"
    ).select_related("persona", "club_destino")
    return render(request, "federacion_app/solicitudes_a_liberar.html", {"solicitudes": solicitudes})


@login_required
@user_passes_test(es_delegado)
def liberar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(
        SolicitudPase, id=solicitud_id, club_origen=request.user.club, estado="pendiente_liberacion"
    )

    if request.method == "POST":
        accion = request.POST.get("accion")
        if accion == "liberar":
            nota = request.POST.get("nota_liberacion", "")
            solicitud.liberar(usuario_libera=request.user, nota=nota)

            archivo_pase = request.FILES.get("formulario_pase_libre")
            if archivo_pase:
                DocumentoSolicitud.objects.create(
                    solicitud=solicitud, archivo=archivo_pase, descripcion="Formulario de pase libre firmado"
                )

            messages.success(request, f"Liberaste a {solicitud.persona}. Ahora la federación revisa el pase.")
        else:
            motivo = request.POST.get("motivo_rechazo", "")
            solicitud.rechazar_liberacion(usuario_rechaza=request.user, motivo=motivo)
            messages.info(request, f"Rechazaste la salida de {solicitud.persona}.")
        return redirect("solicitudes_a_liberar")

    return render(request, "federacion_app/liberar_solicitud.html", {"solicitud": solicitud})


# ---------------------------------------------------------------------
# PANTALLA 2: panel de la federación para revisar/aprobar
# ---------------------------------------------------------------------

@login_required
@user_passes_test(es_federacion)
def panel_solicitudes(request):
    pendientes = SolicitudPase.objects.filter(
        estado__in=["pendiente", "en_revision"]
    ).select_related("persona", "club_origen", "club_destino").order_by("fecha_creacion")
    return render(request, "federacion_app/panel_solicitudes.html", {"solicitudes": pendientes})


@login_required
@user_passes_test(es_federacion)
def historial_solicitudes(request):
    """Historial de todas las solicitudes ya resueltas, de cualquier club."""
    solicitudes = SolicitudPase.objects.filter(
        estado__in=["aprobado", "rechazado"]
    ).select_related("persona", "club_origen", "club_destino").order_by("-fecha_resolucion")
    return render(request, "federacion_app/historial_solicitudes.html", {"solicitudes": solicitudes})


@login_required
@user_passes_test(es_federacion)
def revisar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudPase, id=solicitud_id)

    if request.method == "POST":
        form = ResolucionSolicitudForm(request.POST)
        if form.is_valid():
            accion = form.cleaned_data["accion"]
            if accion == "aprobar":
                solicitud.fecha_resolucion = date.today()
                solicitud.aprobar(usuario_resuelve=request.user)
                messages.success(request, f"Solicitud de {solicitud.persona} aprobada.")
            else:
                solicitud.estado = "rechazado"
                solicitud.motivo_rechazo = form.cleaned_data.get("motivo_rechazo", "")
                solicitud.resuelto_por = request.user
                solicitud.fecha_resolucion = date.today()
                solicitud.save()
                messages.info(request, f"Solicitud de {solicitud.persona} rechazada.")
            return redirect("panel_solicitudes")
    else:
        form = ResolucionSolicitudForm()

    return render(request, "federacion_app/revisar_solicitud.html", {
        "solicitud": solicitud, "form": form,
    })


@login_required
def formulario_09(request, solicitud_id):
    """Genera el Formulario 09: pase entre clubes."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    import io

    solicitud = get_object_or_404(SolicitudPase, id=solicitud_id)
    persona = solicitud.persona

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=estilos["Title"], alignment=TA_CENTER, fontSize=14)
    estilo_normal = ParagraphStyle("normal", parent=estilos["Normal"], fontSize=10, leading=17, alignment=TA_JUSTIFY)
    elementos = []

    logo_federacion = finders.find("federacion_app/logo.png")
    if logo_federacion:
        elementos.append(Image(logo_federacion, width=2.3 * cm, height=2.3 * cm))
        elementos.append(Spacer(1, 6))

    elementos.append(Paragraph("FORM. 09 — Pases Inter Clubes", estilo_titulo))
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph("Sr. Presidente<br/>Federación Fueguina Fútbol de Salón -Futsal-", estilos["Normal"]))
    elementos.append(Spacer(1, 14))

    texto = (
        f"El que suscribe, <b>{persona.apellido}, {persona.nombre}</b>, solicita a las autoridades que correspondan, "
        "me sea concedido según los trámites de rigor el PASE INTER CLUBES como JUGADOR del CLUB "
        f"<b>{solicitud.club_origen.nombre if solicitud.club_origen else '.....'}</b> al CLUB "
        f"<b>{solicitud.club_destino.nombre}</b>."
    )
    elementos.append(Paragraph(texto, estilo_normal))
    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph(f"RÍO GRANDE, ........... de .......................... del {date.today().year}.-", estilos["Normal"]))
    elementos.append(Spacer(1, 24))

    elementos.append(Paragraph(
        f"Firma del jugador: _______________________________&nbsp;&nbsp;&nbsp; "
        f"D.N.I. N°: {persona.documento}&nbsp;&nbsp;&nbsp; CARNET N°: {persona.numero_carnet or '____________'}",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 16))
    elementos.append(Paragraph(
        "NOTA: se requiere la firma del padre/tutor si el jugador solicitante del PASE INTER CLUBES pertenece "
        "a las categorías: C9, C11, C13, C15, C17.",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph(
        "Firma y aclaración padre/tutor: ____________________________&nbsp;&nbsp;&nbsp; "
        "D.N.I. N°: ____________________________",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 26))

    conformidad = [
        [Paragraph("<b>CONFORMIDAD DEL CLUB QUE ABANDONA</b>", estilos["Normal"]),
         Paragraph("<b>CONFORMIDAD DEL CLUB A QUE INGRESA</b>", estilos["Normal"])],
        [Paragraph(f"{solicitud.club_origen.nombre if solicitud.club_origen else ''}", estilos["Normal"]),
         Paragraph(f"{solicitud.club_destino.nombre}", estilos["Normal"])],
        ["", ""],
        ["_______________________&nbsp;&nbsp;&nbsp;SELLO&nbsp;&nbsp;&nbsp;_______________________<br/>firma secretario / firma presidente",
         "_______________________&nbsp;&nbsp;&nbsp;SELLO&nbsp;&nbsp;&nbsp;_______________________<br/>firma secretario / firma presidente"],
    ]
    tabla_conformidad = Table(conformidad, colWidths=[9 * cm, 9 * cm])
    tabla_conformidad.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elementos.append(tabla_conformidad)
    elementos.append(Spacer(1, 20))

    modalidad_texto = solicitud.get_tipo_pase_display() if solicitud.tipo_pase else "DEFINITIVO - PRÉSTAMO (tachar lo que no corresponda)"
    estado_texto = ""
    if solicitud.estado == "aprobado":
        estado_texto = (
            f"Visto la presentación realizada, se resuelve <b>APROBAR</b> el pase solicitado en carácter de "
            f"<b>{modalidad_texto.upper()}</b>. "
            f"Fecha de aprobación: {solicitud.fecha_resolucion.strftime('%d/%m/%Y') if solicitud.fecha_resolucion else ''}"
        )
    else:
        estado_texto = (
            f"Visto la presentación realizada, se resuelve APROBAR el pase solicitado en carácter de "
            f"{modalidad_texto.upper()}."
        )
    elementos.append(Paragraph("<b>Para uso de la FEDERACIÓN o COMISIÓN DE DEPORTES</b>", estilos["Normal"]))
    elementos.append(Spacer(1, 8))
    elementos.append(Paragraph(estado_texto, estilo_normal))
    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph("Firma y sello autoridad actuante: _______________________________________", estilos["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="formulario09_{persona.apellido}.pdf"'
    return response


@login_required
def formulario_10(request, solicitud_id):
    """Genera el Formulario 10: inscripción de jugadores libres o nuevos (habilitación)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    import io

    solicitud = get_object_or_404(SolicitudPase, id=solicitud_id)
    persona = solicitud.persona

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=estilos["Title"], alignment=TA_CENTER, fontSize=14)
    estilo_normal = ParagraphStyle("normal", parent=estilos["Normal"], fontSize=10, leading=17, alignment=TA_JUSTIFY)
    elementos = []

    logo_federacion = finders.find("federacion_app/logo.png")
    if logo_federacion:
        elementos.append(Image(logo_federacion, width=2.3 * cm, height=2.3 * cm))
        elementos.append(Spacer(1, 6))

    elementos.append(Paragraph("FORM. 10 — Inscripción de Jugadores Libres o Nuevos", estilo_titulo))
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph("Sr. Presidente<br/>Federación Fueguina Fútbol de Salón -Futsal-", estilos["Normal"]))
    elementos.append(Spacer(1, 14))

    # LIBRE = ya existía en el sistema sin club; NUEVO = alta recién creada.
    es_nuevo = "NUEVO" if solicitud.tipo == "alta_nueva" else "LIBRE"
    texto = (
        "Se solicita a las autoridades que correspondan, según los trámites de rigor, HABILITAR al siguiente "
        f"jugador <b>{es_nuevo}</b> (tachar lo que no corresponda)."
        "<br/><br/>"
        f"Nombre y Apellido del Jugador: <b>{persona.apellido}, {persona.nombre}</b>"
    )
    elementos.append(Paragraph(texto, estilo_normal))
    elementos.append(Spacer(1, 18))
    elementos.append(Paragraph(
        f"Firma del jugador: _______________________________&nbsp;&nbsp;&nbsp; "
        f"D.N.I. N°: {persona.documento}&nbsp;&nbsp;&nbsp; CARNET N°: {persona.numero_carnet or '____________'}",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 16))
    elementos.append(Paragraph(
        "NOTA: se requiere la firma del padre/tutor si el jugador a HABILITAR pertenece a las categorías: "
        "C9 – C11 – C13 – C15 – C17.",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph(
        "Firma del padre/tutor: ____________________________&nbsp;&nbsp;&nbsp; D.N.I. N°: ____________________________",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 24))

    elementos.append(Paragraph("<b>CONFORMIDAD DEL CLUB A QUE INGRESA</b>", estilos["Normal"]))
    elementos.append(Paragraph(f"{solicitud.club_destino.nombre}", estilos["Normal"]))
    elementos.append(Spacer(1, 26))
    tabla_conformidad = Table(
        [["_______________________&nbsp;&nbsp;&nbsp;SELLO&nbsp;&nbsp;&nbsp;_______________________"],
         ["firma Secretario &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; del CLUB &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; firma Presidente"]],
        colWidths=[18 * cm],
    )
    tabla_conformidad.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 9), ("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    elementos.append(tabla_conformidad)
    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph(f"RÍO GRANDE, ........... de .......................... del {date.today().year}.-", estilos["Normal"]))
    elementos.append(Spacer(1, 24))

    estado_texto = (
        f"Visto la presentación realizada, se resuelve <b>APROBAR</b> la habilitación solicitada. "
        f"Fecha: {solicitud.fecha_resolucion.strftime('%d/%m/%Y') if solicitud.fecha_resolucion else '_______________'}"
        if solicitud.estado == "aprobado" else
        "Visto la presentación realizada, se resuelve APROBAR la habilitación solicitada. Fecha: _______________"
    )
    elementos.append(Paragraph("<b>Para uso de la FEDERACIÓN o COMISIÓN DE DEPORTES</b>", estilos["Normal"]))
    elementos.append(Spacer(1, 8))
    elementos.append(Paragraph(estado_texto, estilo_normal))
    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph("Firma y sello autoridad actuante: _______________________________________", estilos["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="formulario10_{persona.apellido}.pdf"'
    return response


# ---------------------------------------------------------------------
# PANTALLA 3: ficha del jugador con su historial
# ---------------------------------------------------------------------

def _filtrar_padron(request):
    """Lógica de filtro compartida entre la pantalla y las exportaciones."""
    if request.user.rol == "delegado":
        club_id = str(request.user.club_id) if request.user.club_id else None
    else:
        club_id = request.GET.get("club")
    categoria_id = request.GET.get("categoria")
    club_seleccionado = None
    categoria_seleccionada = None
    jugadores = Persona.objects.none()

    if club_id:
        club_seleccionado = get_object_or_404(Club, id=club_id)
        vinculos_activos = Q(vinculos__club=club_seleccionado, vinculos__fecha_fin__isnull=True)

        if categoria_id:
            categoria_seleccionada = get_object_or_404(Categoria, id=categoria_id)
            vinculos_activos &= Q(vinculos__categoria=categoria_seleccionada)

        jugadores = Persona.objects.filter(vinculos_activos).distinct().order_by("apellido")

    return club_seleccionado, categoria_seleccionada, jugadores


@login_required
@user_passes_test(es_federacion)
def importar_excel(request):
    import pandas as pd
    from .forms import ImportarExcelForm

    resultado = None

    if request.method == "POST":
        form = ImportarExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            solo_simular = form.cleaned_data["solo_simular"]

            creados, actualizados = 0, 0
            errores, duplicados = [], []
            vinculos_por_persona = {}
            columnas_esperadas = {"documento", "nombre", "apellido", "fecha_nacimiento"}

            def parsear_fecha(valor):
                if pd.isna(valor):
                    return None
                if hasattr(valor, "date"):
                    return valor.date()
                texto = str(valor).strip()
                for formato in (
                    "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
                    "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                ):
                    try:
                        return datetime.strptime(texto, formato).date()
                    except ValueError:
                        continue
                # Último intento: dejar que pandas adivine el formato.
                try:
                    resultado = pd.to_datetime(texto, dayfirst=True, errors="raise")
                    return resultado.date()
                except (ValueError, TypeError):
                    return None

            # Variantes comunes de nombres de columna que se aceptan como equivalentes.
            alias_columnas = {
                "numero de carnet": "numero_carnet", "número de carnet": "numero_carnet",
                "n° de carnet": "numero_carnet", "nro carnet": "numero_carnet",
                "fecha nacimiento": "fecha_nacimiento", "fecha de nacimiento": "fecha_nacimiento",
                "fecha ingreso": "fecha_ingreso", "fecha de ingreso": "fecha_ingreso",
                "categoría": "categoria",
                "requiere carnet": "requiere_carnet",
            }

            def normalizar_columnas(columnas):
                normalizadas = []
                for c in columnas:
                    c = str(c).strip().lower()
                    c = alias_columnas.get(c, c)
                    normalizadas.append(c)
                return normalizadas

            def leer_hoja_con_encabezado_flexible(archivo, nombre_hoja):
                """
                Busca la fila real de encabezados dentro de las primeras 5 filas,
                por si hay filas vacías o de título arriba de los datos.
                """
                crudo = pd.read_excel(archivo, sheet_name=nombre_hoja, header=None, dtype=str)
                for fila_idx in range(min(5, len(crudo))):
                    candidatas = normalizar_columnas(crudo.iloc[fila_idx].fillna(""))
                    if columnas_esperadas.issubset(set(candidatas)):
                        df = pd.read_excel(archivo, sheet_name=nombre_hoja, header=fila_idx, dtype=str)
                        df.columns = normalizar_columnas(df.columns)
                        return df
                # No se encontró una fila de encabezados válida: devuelve la lectura
                # normal (fila 0) para que el mensaje de error liste qué faltó.
                df = pd.read_excel(archivo, sheet_name=nombre_hoja, dtype=str)
                df.columns = normalizar_columnas(df.columns)
                return df

            try:
                excel = pd.ExcelFile(archivo)
            except Exception as e:
                errores.append(f"No se pudo leer el archivo: {e}")
                excel = None

            if excel:
                for nombre_hoja in excel.sheet_names:
                    nombre_club = nombre_hoja.strip()
                    df = leer_hoja_con_encabezado_flexible(archivo, nombre_hoja)

                    faltantes = columnas_esperadas - set(df.columns)
                    if faltantes:
                        errores.append(f"Hoja '{nombre_hoja}': faltan columnas {faltantes}, se omite.")
                        continue

                    club_de_la_hoja, _ = Club.objects.get_or_create(
                        nombre__iexact=nombre_club,
                        defaults={
                            "nombre": nombre_club,
                            "codigo_afiliacion": f"AUTO-{nombre_club[:10].upper()}",
                            "fecha_afiliacion": date.today(),
                        },
                    )

                    for i, fila in df.iterrows():
                        ubicacion = f"Hoja '{nombre_hoja}', fila {i + 2}"
                        try:
                            documento = str(fila["documento"]).strip()
                            if not documento or documento.lower() == "nan":
                                errores.append(f"{ubicacion}: sin documento, se omite")
                                continue

                            # Si la fila trae su propia columna 'club', se usa esa
                            # en vez del nombre de la hoja (útil cuando el archivo
                            # tiene una sola hoja con varios clubes mezclados).
                            if "club" in df.columns and pd.notna(fila.get("club")) and str(fila["club"]).strip():
                                nombre_club_fila = str(fila["club"]).strip()
                                club, _ = Club.objects.get_or_create(
                                    nombre__iexact=nombre_club_fila,
                                    defaults={
                                        "nombre": nombre_club_fila,
                                        "codigo_afiliacion": f"AUTO-{nombre_club_fila[:10].upper()}",
                                        "fecha_afiliacion": date.today(),
                                    },
                                )
                            else:
                                club = club_de_la_hoja

                            fecha_nac = parsear_fecha(fila["fecha_nacimiento"])
                            if not fecha_nac:
                                errores.append(f"{ubicacion}: fecha de nacimiento inválida")
                                continue

                            categoria = None
                            if "categoria" in df.columns and pd.notna(fila.get("categoria")):
                                categoria, _ = Categoria.objects.get_or_create(
                                    nombre__iexact=str(fila["categoria"]).strip(),
                                    defaults={"nombre": str(fila["categoria"]).strip()},
                                )

                            fecha_inicio = date.today()
                            if "fecha_ingreso" in df.columns:
                                f = parsear_fecha(fila.get("fecha_ingreso"))
                                if f:
                                    fecha_inicio = f

                            numero_carnet = ""
                            if "numero_carnet" in df.columns and pd.notna(fila.get("numero_carnet")):
                                numero_carnet = str(fila["numero_carnet"]).strip()

                            requiere_carnet = False
                            if "requiere_carnet" in df.columns and pd.notna(fila.get("requiere_carnet")):
                                requiere_carnet = str(fila["requiere_carnet"]).strip().lower() in ("si", "sí", "1", "true", "x")

                            if not solo_simular:
                                persona, fue_creada = Persona.objects.get_or_create(
                                    documento=documento,
                                    defaults={
                                        "tipo": "jugador",
                                        "nombre": str(fila["nombre"]).strip(),
                                        "apellido": str(fila["apellido"]).strip(),
                                        "fecha_nacimiento": fecha_nac,
                                        "numero_carnet": numero_carnet,
                                        "requiere_carnet": requiere_carnet,
                                    },
                                )
                                if fue_creada:
                                    creados += 1
                                else:
                                    actualizados += 1
                                    duplicados.append(f"{ubicacion}: {documento} aparece en más de una hoja")
                                vinculos_por_persona.setdefault(documento, []).append(
                                    (persona, club, categoria, fecha_inicio)
                                )
                            else:
                                # En modo simulación solo contamos, sin tocar la base.
                                if Persona.objects.filter(documento=documento).exists():
                                    actualizados += 1
                                    duplicados.append(f"{ubicacion}: {documento} aparece en más de una hoja")
                                else:
                                    creados += 1

                        except Exception as e:
                            errores.append(f"{ubicacion}: error inesperado — {e}")

                if not solo_simular:
                    for documento, lista in vinculos_por_persona.items():
                        lista.sort(key=lambda v: v[3])
                        for idx, (persona, club, categoria, fecha_inicio) in enumerate(lista):
                            es_el_ultimo = idx == len(lista) - 1
                            fecha_fin = None if es_el_ultimo else lista[idx + 1][3]

                            if not persona.vinculos.filter(club=club, fecha_inicio=fecha_inicio).exists():
                                if fecha_fin is None:
                                    # Va a quedar como vínculo activo: hay que cerrar
                                    # cualquier otro vínculo activo que ya tuviera
                                    # (de una carga anterior), para no violar la regla
                                    # de "un solo club activo por persona".
                                    persona.vinculos.filter(fecha_fin__isnull=True).exclude(
                                        club=club, fecha_inicio=fecha_inicio
                                    ).update(fecha_fin=fecha_inicio)

                                Vinculo.objects.create(
                                    persona=persona, club=club, categoria=categoria,
                                    fecha_inicio=fecha_inicio, fecha_fin=fecha_fin,
                                )

            resultado = {
                "simulado": solo_simular,
                "creados": creados, "actualizados": actualizados,
                "errores": errores, "duplicados": duplicados,
            }
            if not solo_simular:
                messages.success(request, f"Importación completa: {creados} jugadores nuevos.")
    else:
        form = ImportarExcelForm()

    return render(request, "federacion_app/importar_excel.html", {"form": form, "resultado": resultado})


@login_required
@user_passes_test(es_delegado)
def inscribir_torneo(request):
    from .models import InscripcionTorneo, Torneo

    torneos = Torneo.objects.filter(activo=True)
    categorias = Categoria.objects.all().order_by("nombre")
    categoria_id = request.GET.get("categoria") or request.POST.get("categoria_filtro")

    jugadores_qs = Persona.objects.filter(
        vinculos__club=request.user.club, vinculos__fecha_fin__isnull=True
    ).distinct().order_by("apellido")
    if categoria_id:
        jugadores_qs = jugadores_qs.filter(vinculos__categoria_id=categoria_id, vinculos__fecha_fin__isnull=True)

    # Para marcar en la tabla a quiénes ya se inscribió a cada torneo, sin
    # tener que abrir cada uno: se arma un set de (persona_id, torneo_id).
    ya_inscritos = set(
        InscripcionTorneo.objects.filter(club=request.user.club)
        .values_list("persona_id", "torneo_id")
    )

    if request.method == "POST":
        torneo_id = request.POST.get("torneo")
        ids_seleccionados = request.POST.getlist("jugadores")
        torneo = get_object_or_404(Torneo, id=torneo_id) if torneo_id else None

        if not torneo:
            messages.error(request, "Elegí un torneo antes de inscribir.")
        elif not ids_seleccionados:
            messages.error(request, "Marcá al menos un jugador para inscribir.")
        else:
            creadas = 0
            for persona_id in ids_seleccionados:
                if (int(persona_id), torneo.id) in ya_inscritos:
                    continue  # ya estaba inscrito a ese torneo, se salta
                InscripcionTorneo.objects.create(
                    persona_id=persona_id, club=request.user.club,
                    torneo=torneo, inscrito_por=request.user,
                )
                creadas += 1
            messages.success(request, f"Inscribiste {creadas} jugador(es) a {torneo}. La federación va a liquidar el costo.")
            return redirect("inscribir_torneo")

    mis_inscripciones = InscripcionTorneo.objects.filter(club=request.user.club).select_related("persona", "torneo")
    return render(request, "federacion_app/inscribir_torneo.html", {
        "torneos": torneos, "categorias": categorias, "categoria_id": categoria_id,
        "jugadores": jugadores_qs, "ya_inscritos": ya_inscritos,
        "inscripciones": mis_inscripciones,
    })


@login_required
@user_passes_test(es_delegado)
def mis_inscripciones_torneo(request):
    """Historial de todas las inscripciones a torneos hechas por el club del delegado."""
    from .models import InscripcionTorneo
    inscripciones = InscripcionTorneo.objects.filter(
        club=request.user.club
    ).select_related("persona", "torneo").order_by("-fecha_inscripcion")
    return render(request, "federacion_app/mis_inscripciones_torneo.html", {"inscripciones": inscripciones})


@login_required
@user_passes_test(es_federacion)
def marcar_pagado(request, inscripcion_id):
    from .models import InscripcionTorneo
    inscripcion = get_object_or_404(InscripcionTorneo, id=inscripcion_id)
    if request.method == "POST":
        inscripcion.pagado = not inscripcion.pagado
        inscripcion.fecha_pago = date.today() if inscripcion.pagado else None
        inscripcion.save()
    return redirect("cobro_masivo")


@login_required
@user_passes_test(es_federacion)
def historial_inscripciones(request):
    from .models import InscripcionTorneo
    inscripciones = InscripcionTorneo.objects.filter(pagado=True).select_related(
        "persona", "club", "torneo"
    ).order_by("-fecha_pago")
    return render(request, "federacion_app/historial_inscripciones.html", {"inscripciones": inscripciones})


@login_required
@user_passes_test(es_federacion)
def cobro_masivo(request):
    """
    Pantalla única de cobro: arriba los totales pendientes por club y
    categoría (de todos los torneos); al elegir un torneo (y opcionalmente
    una categoría), aparece la grilla para tildar los tres conceptos y
    validar el pago. Al validar el pago, el jugador desaparece de la lista.
    """
    from collections import defaultdict
    from .models import InscripcionTorneo, Torneo

    torneos = Torneo.objects.filter(activo=True)
    categorias = Categoria.objects.all().order_by("nombre")
    clubes = Club.objects.all().order_by("nombre")

    # --- Totales pendientes (de todos los torneos, no solo el filtrado) ---
    pendientes_generales = InscripcionTorneo.objects.filter(
        estado="cobrado", pagado=False
    ).select_related("persona", "club")

    totales_club = defaultdict(float)
    totales_categoria = defaultdict(float)
    for i in pendientes_generales:
        monto = float(i.monto_total)
        totales_club[i.club.nombre] += monto
        vinculo = i.persona.vinculos.filter(club=i.club, fecha_fin__isnull=True).first()
        nombre_categoria = vinculo.categoria.nombre if vinculo and vinculo.categoria else "Sin categoría"
        totales_categoria[nombre_categoria] += monto

    total_pagado = sum(
        float(i.monto_total) for i in InscripcionTorneo.objects.filter(pagado=True)
    )

    # --- Grilla filtrada por torneo / categoría / club ---
    torneo_id = request.POST.get("torneo") or request.GET.get("torneo")
    categoria_id = request.POST.get("categoria") or request.GET.get("categoria")
    club_id = request.POST.get("club") or request.GET.get("club")
    torneo = get_object_or_404(Torneo, id=torneo_id) if torneo_id else None

    filas = []
    if torneo:
        qs = InscripcionTorneo.objects.filter(torneo=torneo, pagado=False).select_related("persona", "club")
        if club_id:
            qs = qs.filter(club_id=club_id)
        for i in qs:
            vinculo = i.persona.vinculos.filter(club=i.club, fecha_fin__isnull=True).first()
            cat = vinculo.categoria if vinculo else None
            if categoria_id and (not cat or str(cat.id) != categoria_id):
                continue
            filas.append({
                "inscripcion": i,
                "der_marcado": i.derechos_federativos is not None,
                "fondo_marcado": i.fondo_seleccion is not None,
                "carnet_marcado": i.carnet is not None,
                "jugador_libre_marcado": i.jugador_libre is not None,
                "fichaje_nuevo_marcado": i.fichaje_nuevo is not None,
            })

    if request.method == "POST" and torneo:
        cambios, pagos = 0, 0
        for fila in filas:
            i = fila["inscripcion"]
            tocado = False
            if request.POST.get(f"der_{i.id}"):
                i.derechos_federativos = torneo.precio_derechos_federativos
                tocado = True
            if request.POST.get(f"fondo_{i.id}"):
                i.fondo_seleccion = torneo.precio_fondo_seleccion
                tocado = True
            if request.POST.get(f"carnet_{i.id}"):
                i.carnet = torneo.precio_carnet
                tocado = True
            if request.POST.get(f"jugadorlibre_{i.id}"):
                i.jugador_libre = torneo.precio_jugador_libre
                tocado = True
            if request.POST.get(f"fichajenuevo_{i.id}"):
                i.fichaje_nuevo = torneo.precio_fichaje_nuevo
                tocado = True
            if tocado:
                i.estado = "cobrado"
                i.cobrado_por = request.user
                i.fecha_cobro = date.today()
                cambios += 1

            # Validar el pago: si se tilda, el jugador desaparece de la
            # lista (queda excluido por el filtro pagado=False de arriba).
            if request.POST.get(f"validar_{i.id}"):
                i.pagado = True
                i.fecha_pago = date.today()
                pagos += 1

            if tocado or request.POST.get(f"validar_{i.id}"):
                i.save()

        if cambios or pagos:
            partes = []
            if cambios:
                partes.append(f"{cambios} cobro(s) registrado(s)")
            if pagos:
                partes.append(f"{pagos} pago(s) validado(s)")
            messages.success(request, " y ".join(partes) + ".")

        url = f"/torneos/cobro-masivo/?torneo={torneo.id}"
        if categoria_id:
            url += f"&categoria={categoria_id}"
        if club_id:
            url += f"&club={club_id}"
        return redirect(url)

    return render(request, "federacion_app/cobro_masivo.html", {
        "torneos": torneos, "categorias": categorias, "clubes": clubes,
        "torneo": torneo, "categoria_id": categoria_id, "club_id": club_id, "filas": filas,
        "totales_club": sorted(totales_club.items()),
        "totales_categoria": sorted(totales_categoria.items()),
        "total_general": sum(totales_club.values()),
        "total_pagado": total_pagado,
    })


@login_required
@user_passes_test(es_federacion)
def cobrar_inscripcion(request, inscripcion_id):
    from .models import InscripcionTorneo
    from .forms import CobroInscripcionForm

    inscripcion = get_object_or_404(InscripcionTorneo, id=inscripcion_id)

    if request.method == "POST":
        form = CobroInscripcionForm(request.POST)
        if form.is_valid():
            inscripcion.derechos_federativos = form.cleaned_data["derechos_federativos"]
            inscripcion.fondo_seleccion = form.cleaned_data["fondo_seleccion"]
            inscripcion.carnet = form.cleaned_data["carnet"]
            inscripcion.jugador_libre = form.cleaned_data["jugador_libre"]
            inscripcion.fichaje_nuevo = form.cleaned_data["fichaje_nuevo"]
            inscripcion.estado = "cobrado"
            inscripcion.cobrado_por = request.user
            inscripcion.fecha_cobro = date.today()
            inscripcion.save()
            messages.success(request, f"Cobro registrado para {inscripcion.persona}.")
            return redirect("cobro_masivo")
    else:
        form = CobroInscripcionForm(initial={
            "derechos_federativos": inscripcion.derechos_federativos,
            "fondo_seleccion": inscripcion.fondo_seleccion,
            "carnet": inscripcion.carnet,
            "jugador_libre": inscripcion.jugador_libre,
            "fichaje_nuevo": inscripcion.fichaje_nuevo,
        })

    return render(request, "federacion_app/cobrar_inscripcion.html", {
        "inscripcion": inscripcion, "form": form,
    })


def _ids_jugadores_sancionados(club):
    """
    IDs de jugadores del club que hoy no pueden jugar: sanción por
    tarjetas todavía pendiente de resolver, o sanción disciplinaria activa.
    """
    from .models import SancionTarjeta, SancionDisciplinaria
    ids_tarjetas = set(
        SancionTarjeta.objects.filter(club=club, estado="pendiente").values_list("persona_id", flat=True)
    )
    ids_disciplinarias = set(
        SancionDisciplinaria.objects.filter(club=club, estado="activa").values_list("persona_id", flat=True)
    )
    return ids_tarjetas | ids_disciplinarias


@login_required
@user_passes_test(es_delegado)
def planilla_partido(request):
    """
    Genera la planilla de buena fe (PDF) para un partido: el delegado
    tilda los jugadores de su plantel y les asigna número de camiseta,
    completa los datos del partido, y descarga el PDF listo para imprimir.
    """
    club = request.user.club
    jugadores_club = Persona.objects.filter(
        vinculos__club=club, vinculos__fecha_fin__isnull=True, activo=True
    ).distinct().order_by("apellido")
    categorias = Categoria.objects.all().order_by("nombre")

    ids_sancionados = _ids_jugadores_sancionados(club)

    # Categoría de cada jugador, para el filtro del lado del cliente
    # (no recarga la página, así no se pierden los tildados al cambiar filtro).
    # Los sancionados no entran a la lista seleccionable.
    jugadores_con_categoria = []
    jugadores_sancionados = []
    for p in jugadores_club:
        vinculo = p.vinculos.filter(club=club, fecha_fin__isnull=True).first()
        if p.id in ids_sancionados:
            jugadores_sancionados.append(p)
            continue
        jugadores_con_categoria.append({
            "persona": p,
            "categoria_id": vinculo.categoria_id if vinculo and vinculo.categoria_id else "",
        })

    if request.method == "POST":
        return _generar_pdf_planilla(request, club)

    return render(request, "federacion_app/planilla_partido.html", {
        "jugadores": jugadores_con_categoria, "categorias": categorias, "club": club,
        "jugadores_sancionados": jugadores_sancionados,
    })


def _generar_pdf_planilla(request, club):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=estilos["Title"], alignment=TA_CENTER, fontSize=20)
    elementos = []

    # --- Encabezado: logo federación | nombre del club | escudo del club ---
    logo_federacion = finders.find("federacion_app/logo.png")
    img_federacion = Image(logo_federacion, width=2.5 * cm, height=2.5 * cm) if logo_federacion else ""
    img_club = Image(club.escudo.path, width=2.5 * cm, height=2.5 * cm) if club.escudo else ""

    encabezado = Table(
        [[img_federacion, Paragraph(club.nombre.upper(), estilo_titulo), img_club]],
        colWidths=[3.5 * cm, 11 * cm, 3.5 * cm],
    )
    encabezado.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    elementos.append(encabezado)
    elementos.append(Spacer(1, 16))

    # --- Datos del partido ---
    datos = [
        [f"Torneo: {request.POST.get('torneo', '')}", f"Categoria: {request.POST.get('categoria', '')}"],
        [f"Lugar: {request.POST.get('lugar', '')}", f"Fecha: {request.POST.get('fecha', '')}"],
        [f"Delegada/o: {request.POST.get('delegado', '')}", ""],
    ]
    tabla_datos = Table(datos, colWidths=[10 * cm, 8 * cm])
    tabla_datos.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("SPAN", (0, 2), (1, 2)),
    ]))
    elementos.append(tabla_datos)
    elementos.append(Spacer(1, 14))

    # --- Plantel: hasta 15 jugadores tildados por el delegado + DT/AT/PF ---
    filas = [["", "APELLIDO Y NOMBRE", "CARNET/DNI", "N° CAMISETA"]]
    ids_seleccionados = request.POST.getlist("jugadores")
    ids_sancionados = _ids_jugadores_sancionados(club)
    ids_seleccionados = [pid for pid in ids_seleccionados if int(pid) not in ids_sancionados]
    seleccionados = list(Persona.objects.filter(id__in=ids_seleccionados))
    # Mantiene el orden en que vinieron tildados en el formulario
    orden = {int(pid): i for i, pid in enumerate(ids_seleccionados)}
    seleccionados.sort(key=lambda p: orden.get(p.id, 999))

    for i in range(1, 16):
        if i <= len(seleccionados):
            p = seleccionados[i - 1]
            nombre_completo = f"{p.apellido}, {p.nombre}"
            carnet = p.numero_carnet or p.documento
            camiseta = request.POST.get(f"camiseta_{p.id}", "")
        else:
            nombre_completo, carnet, camiseta = "", "", ""
        filas.append([str(i), nombre_completo, carnet, camiseta])

    filas.append(["DT", request.POST.get("dt", ""), "", ""])
    filas.append(["AT", request.POST.get("at", ""), "", ""])
    filas.append(["PF", request.POST.get("pf", ""), "", ""])

    tabla_plantel = Table(filas, colWidths=[1.3 * cm, 9 * cm, 4 * cm, 3.7 * cm], repeatRows=1)
    tabla_plantel.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e5e7eb")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabla_plantel)
    elementos.append(Spacer(1, 14))

    # --- Hora de entrega ---
    tabla_hora = Table([[f"HORA DE ENTREGA: {request.POST.get('hora_entrega', '')}"]], colWidths=[18 * cm])
    tabla_hora.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.7, colors.black), ("TOPPADDING", (0, 0), (-1, -1), 5)]))
    elementos.append(tabla_hora)
    elementos.append(Spacer(1, 14))

    # --- Arbitraje / Mesa ---
    pagos = [
        ["", "EFECTIVO", "TRANSFERENCIA", "", "EFECTIVO", "TRANSFERENCIA"],
        ["ARBITRAJE", request.POST.get("arbitraje_efectivo", ""), request.POST.get("arbitraje_transferencia", ""),
         "MESA", request.POST.get("mesa_efectivo", ""), request.POST.get("mesa_transferencia", "")],
    ]
    tabla_pagos = Table(pagos, colWidths=[3 * cm, 3.5 * cm, 3.5 * cm, 2.5 * cm, 2.7 * cm, 2.8 * cm])
    tabla_pagos.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elementos.append(tabla_pagos)

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="planilla_{club.nombre}.pdf"'
    return response


@login_required
@user_passes_test(es_delegado)
def mi_padron(request):
    """El delegado ve el padrón completo de su propio club, filtrable por categoría."""
    club = request.user.club
    categorias = Categoria.objects.all().order_by("nombre")
    categoria_id = request.GET.get("categoria")

    jugadores = Persona.objects.filter(
        vinculos__club=club, vinculos__fecha_fin__isnull=True, tipo="jugador"
    ).distinct().order_by("apellido")
    if categoria_id:
        jugadores = jugadores.filter(vinculos__categoria_id=categoria_id, vinculos__fecha_fin__isnull=True)

    return render(request, "federacion_app/mi_padron.html", {
        "club": club, "categorias": categorias, "categoria_id": categoria_id, "jugadores": jugadores,
    })


@login_required
@user_passes_test(es_federacion)
def padron_club(request):
    """Lista los jugadores activos de un club elegido, con filtro opcional por categoría. Solo federación."""
    clubes = Club.objects.all().order_by("nombre")
    categorias = Categoria.objects.all().order_by("nombre")
    club_seleccionado, categoria_seleccionada, jugadores = _filtrar_padron(request)

    return render(request, "federacion_app/padron_club.html", {
        "clubes": clubes, "categorias": categorias,
        "club_seleccionado": club_seleccionado, "categoria_seleccionada": categoria_seleccionada,
        "jugadores": jugadores,
    })


@login_required
@user_passes_test(lambda u: u.rol in ("federacion", "delegado"))
def padron_excel(request):
    import openpyxl
    from openpyxl.styles import Font
    from django.http import HttpResponse

    club_seleccionado, categoria_seleccionada, jugadores = _filtrar_padron(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Padrón"

    titulo = f"Padrón — {club_seleccionado.nombre if club_seleccionado else 'todos los clubes'}"
    if categoria_seleccionada:
        titulo += f" — {categoria_seleccionada.nombre}"
    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])

    ws.append(["Apellido", "Nombre", "Documento", "Categoría", "Desde"])
    for celda in ws[3]:
        celda.font = Font(bold=True)

    for p in jugadores:
        vinculo = p.vinculos.filter(club=club_seleccionado, fecha_fin__isnull=True).first()
        ws.append([
            p.apellido, p.nombre, p.documento,
            vinculo.categoria.nombre if vinculo and vinculo.categoria else "",
            vinculo.fecha_inicio.strftime("%d/%m/%Y") if vinculo else "",
        ])

    for col in ws.columns:
        largo = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = largo + 4

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    nombre_archivo = f"padron_{club_seleccionado.nombre if club_seleccionado else 'general'}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.rol in ("federacion", "delegado"))
def padron_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse
    import io

    club_seleccionado, categoria_seleccionada, jugadores = _filtrar_padron(request)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    estilos = getSampleStyleSheet()
    elementos = []

    titulo = f"Padrón — {club_seleccionado.nombre if club_seleccionado else 'todos los clubes'}"
    if categoria_seleccionada:
        titulo += f" — {categoria_seleccionada.nombre}"
    elementos.append(Paragraph(titulo, estilos["Title"]))
    elementos.append(Spacer(1, 12))

    datos = [["Apellido", "Nombre", "Documento", "Categoría", "Desde"]]
    for p in jugadores:
        vinculo = p.vinculos.filter(club=club_seleccionado, fecha_fin__isnull=True).first()
        datos.append([
            p.apellido, p.nombre, p.documento,
            vinculo.categoria.nombre if vinculo and vinculo.categoria else "—",
            vinculo.fecha_inicio.strftime("%d/%m/%Y") if vinculo else "—",
        ])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fb")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    nombre_archivo = f"padron_{club_seleccionado.nombre if club_seleccionado else 'general'}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
def ficha_persona(request, persona_id):
    persona = get_object_or_404(Persona, id=persona_id)
    historial = persona.vinculos.select_related("club", "categoria").order_by("-fecha_inicio")

    edad = None
    if persona.fecha_nacimiento:
        hoy = date.today()
        edad = hoy.year - persona.fecha_nacimiento.year - (
            (hoy.month, hoy.day) < (persona.fecha_nacimiento.month, persona.fecha_nacimiento.day)
        )

    autorizacion = persona.documentos.filter(tipo="autorizacion_fichaje").first()
    tarjetas = persona.tarjetas.select_related("club", "torneo").order_by("-fecha_partido")
    sanciones = persona.sanciones.order_by("-fecha_generada")
    sanciones_disciplinarias = persona.sanciones_disciplinarias.order_by("-fecha_sancion")
    clubes_activos = persona.vinculos_activos.order_by("-es_principal")

    # El delegado puede subir/cambiar la foto solo de la gente de su propio club.
    puede_editar_foto = (
        request.user.rol == "delegado" and request.user.club and
        persona.vinculos.filter(club=request.user.club, fecha_fin__isnull=True).exists()
    )

    # Si es un delegado y el jugador es suyo, guardamos la categoría para armar
    # el acceso directo al Formulario 12 (ya con torneo/categoría precargados).
    categoria_para_formulario_12 = None
    if puede_editar_foto and persona.tipo == "jugador":
        vinculo_propio = persona.vinculos.filter(club=request.user.club, fecha_fin__isnull=True).first()
        if vinculo_propio:
            categoria_para_formulario_12 = vinculo_propio.categoria

    # Si el delegado ve a un jugador que todavía no es suyo, le damos un
    # acceso directo para pedirlo (pase), sin tener que anotar el documento.
    puede_pedir_jugador = (
        request.user.rol == "delegado" and request.user.club and persona.tipo == "jugador" and
        not puede_editar_foto
    )

    return render(request, "federacion_app/ficha_persona.html", {
        "persona": persona, "historial": historial, "edad": edad, "autorizacion": autorizacion,
        "tarjetas": tarjetas, "sanciones": sanciones, "sanciones_disciplinarias": sanciones_disciplinarias,
        "puede_editar_foto": puede_editar_foto, "clubes_activos": clubes_activos,
        "categoria_para_formulario_12": categoria_para_formulario_12,
        "puede_pedir_jugador": puede_pedir_jugador,
    })


@login_required
def subir_foto_persona(request, persona_id):
    """El delegado sube o cambia la foto de un jugador o técnico de su propio club."""
    persona = get_object_or_404(Persona, id=persona_id)
    tiene_permiso = (
        request.user.rol == "delegado" and request.user.club and
        persona.vinculos.filter(club=request.user.club, fecha_fin__isnull=True).exists()
    ) or request.user.rol == "federacion"

    if not tiene_permiso:
        messages.error(request, "No tenés permiso para editar la foto de esta persona.")
        return redirect("ficha_persona", persona_id=persona.id)

    if request.method == "POST" and request.FILES.get("foto"):
        persona.foto = request.FILES["foto"]
        persona.save()
        messages.success(request, "Foto actualizada.")
    return redirect("ficha_persona", persona_id=persona.id)


@login_required
@user_passes_test(es_federacion)
def alternar_activo_persona(request, persona_id):
    """La federación activa o desactiva a un jugador (no puede ser seleccionado en planillas si está inactivo)."""
    persona = get_object_or_404(Persona, id=persona_id)
    if request.method == "POST":
        persona.activo = not persona.activo
        persona.save()
        estado = "activado" if persona.activo else "desactivado"
        messages.success(request, f"{persona} fue {estado}.")
    return redirect("ficha_persona", persona_id=persona.id)


@login_required
@user_passes_test(es_federacion)
def imprimir_carnet(request, persona_id):
    """Genera en PDF el carnet provisorio del jugador, con el mismo diseño que se ve en la ficha."""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor, white
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    import io

    persona = get_object_or_404(Persona, id=persona_id)

    edad = None
    if persona.fecha_nacimiento:
        hoy = date.today()
        edad = hoy.year - persona.fecha_nacimiento.year - (
            (hoy.month, hoy.day) < (persona.fecha_nacimiento.month, persona.fecha_nacimiento.day)
        )

    NARANJA = HexColor("#f97316")
    NARANJA_OSCURO = HexColor("#ea580c")
    NAVY = HexColor("#0f2a4a")
    NAVY_CLARO = HexColor("#16385e")
    GRIS = HexColor("#374151")

    buffer = io.BytesIO()
    ancho_pagina, alto_pagina = A4
    c = canvas.Canvas(buffer, pagesize=A4)

    # --- Tamaño y posición de la tarjeta, centrada arriba de la hoja ---
    ancho_carnet = 17 * cm
    alto_carnet = 8.6 * cm
    x0 = (ancho_pagina - ancho_carnet) / 2
    y0 = alto_pagina - 3 * cm - alto_carnet

    radio = 0.5 * cm

    # Fondo blanco redondeado
    c.setFillColor(white)
    c.roundRect(x0, y0, ancho_carnet, alto_carnet, radio, fill=1, stroke=0)

    # --- Franja naranja lateral (fondo, sin el ícono/texto todavía) ---
    ancho_franja = 1.6 * cm
    c.saveState()
    p = c.beginPath()
    p.moveTo(x0 + radio, y0)
    p.lineTo(x0 + ancho_franja, y0)
    p.lineTo(x0 + ancho_franja, y0 + alto_carnet)
    p.lineTo(x0 + radio, y0 + alto_carnet)
    p.arcTo(x0, y0 + alto_carnet - 2 * radio, x0 + 2 * radio, y0 + alto_carnet, startAng=90, extent=90)
    p.lineTo(x0, y0 + radio)
    p.arcTo(x0, y0, x0 + 2 * radio, y0 + 2 * radio, startAng=180, extent=90)
    p.close()
    c.setFillColor(NARANJA)
    c.drawPath(p, fill=1, stroke=0)
    c.restoreState()

    # --- Contenido principal ---
    x_contenido = x0 + ancho_franja + 0.7 * cm
    y_cursor = y0 + alto_carnet - 1.1 * cm

    tipo_texto = "JUGADOR" if persona.tipo == "jugador" else (persona.get_rol_tecnico_display() or "CUERPO TÉCNICO").upper()
    c.setFillColor(NARANJA)
    c.setFont("Helvetica-Bold", 15)
    c.drawString(x_contenido, y_cursor, tipo_texto)

    y_cursor -= 0.65 * cm
    c.setFillColor(HexColor("#111827"))
    c.setFont("Helvetica-Bold", 15)
    c.drawString(x_contenido, y_cursor, f"{persona.apellido} {persona.nombre}")

    y_cursor -= 0.5 * cm
    c.setFillColor(GRIS)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x_contenido, y_cursor, f"{persona.documento} DNI")

    if persona.fecha_nacimiento:
        y_cursor -= 0.4 * cm
        texto_fecha = persona.fecha_nacimiento.strftime("%d/%m/%Y")
        if edad is not None:
            texto_fecha += f" · {edad} años"
        c.drawString(x_contenido, y_cursor, texto_fecha)

    y_cursor -= 0.45 * cm
    c.setFont("Helvetica-Bold", 10)
    club_texto = f"Club actual: {persona.club_actual.nombre}" if persona.club_actual else "Sin club"
    c.drawString(x_contenido, y_cursor, club_texto)

    # Foto arriba a la derecha
    ancho_foto, alto_foto = 3.3 * cm, 3.9 * cm
    x_foto = x0 + ancho_carnet - ancho_foto - 0.7 * cm
    y_foto = y0 + alto_carnet - alto_foto - 0.6 * cm
    if persona.foto and hasattr(persona.foto, "path"):
        try:
            c.drawImage(persona.foto.path, x_foto, y_foto, width=ancho_foto, height=alto_foto,
                        preserveAspectRatio=False, mask="auto")
        except Exception:
            pass
    else:
        c.setFillColor(HexColor("#f4f6f9"))
        c.rect(x_foto, y_foto, ancho_foto, alto_foto, fill=1, stroke=1)
        c.setFillColor(HexColor("#9ca3af"))
        c.setFont("Helvetica-Bold", 22)
        iniciales = f"{persona.nombre[:1]}{persona.apellido[:1]}".upper()
        c.drawCentredString(x_foto + ancho_foto / 2, y_foto + alto_foto / 2 - 8, iniciales)

    # --- Franja azul diagonal, abajo ---
    c.saveState()
    p2 = c.beginPath()
    alto_azul_izq = alto_carnet * 0.26
    alto_azul_der = alto_carnet * 0.40
    p2.moveTo(x0, y0)
    p2.lineTo(x0 + ancho_carnet - radio, y0)
    p2.arcTo(x0 + ancho_carnet - 2 * radio, y0, x0 + ancho_carnet, y0 + 2 * radio, startAng=270, extent=90)
    p2.lineTo(x0 + ancho_carnet, y0 + alto_azul_der)
    p2.lineTo(x0, y0 + alto_azul_izq)
    p2.lineTo(x0, y0 + radio)
    p2.arcTo(x0, y0, x0 + 2 * radio, y0 + 2 * radio, startAng=180, extent=90)
    p2.close()
    c.setFillColor(NAVY)
    c.drawPath(p2, fill=1, stroke=0)
    c.restoreState()

    # Texto federación, en blanco, sobre la franja azul
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(x_contenido, y0 + 0.85 * cm, "FEDERACIÓN FUEGUINA")
    c.drawString(x_contenido, y0 + 0.4 * cm, "DE FÚTBOL DE SALÓN")

    # N° de carnet, centrado en la franja azul
    cx_carnet = x0 + ancho_carnet * 0.62
    c.setFillColor(HexColor("#a9c2de"))
    c.setFont("Helvetica-Bold", 7)
    c.drawCentredString(cx_carnet, y0 + 0.95 * cm, "N° DE CARNET")
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 17)
    c.drawCentredString(cx_carnet, y0 + 0.45 * cm, persona.numero_carnet or "—")

    # Logo de la federación, dentro de la franja azul
    logo_path = finders.find("federacion_app/logo.png")
    if logo_path:
        try:
            c.drawImage(logo_path, x0 + ancho_carnet - 1.9 * cm, y0 + 0.35 * cm,
                        width=1.4 * cm, height=1.4 * cm, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    # --- Ícono y texto vertical de la franja naranja, AL FINAL: así quedan
    # por encima del corte diagonal azul y no se tapan (la franja azul avanza
    # un poco dentro del ancho de la franja naranja en su parte más angosta). ---
    cx_icono = x0 + ancho_franja / 2
    cy_icono = y0 + alto_carnet - 1.1 * cm
    c.setFillColor(white)
    c.setFillAlpha(0.3)
    c.circle(cx_icono, cy_icono, 0.42 * cm, fill=1, stroke=0)
    c.setFillAlpha(1)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(cx_icono, cy_icono - 5, "›")

    c.saveState()
    c.translate(x0 + ancho_franja / 2 + 4, y0 + alto_carnet / 2)
    c.rotate(90)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(0, 0, "CREDENCIAL OFICIAL")
    c.restoreState()

    # Contorno de corte y aclaración
    c.setStrokeColor(HexColor("#cbd5e1"))
    c.setDash(3, 3)
    c.roundRect(x0, y0, ancho_carnet, alto_carnet, radio, fill=0, stroke=1)
    c.setDash()
    c.setFillColor(HexColor("#6b7280"))
    c.setFont("Helvetica", 9)
    c.drawCentredString(ancho_pagina / 2, y0 - 0.8 * cm, "Carnet provisorio — recortar por la línea punteada")

    c.showPage()
    c.save()
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="carnet_{persona.apellido}.pdf"'
    return response


@login_required
def subir_autorizacion(request, persona_id):
    """El delegado sube el Formulario 08 ya firmado por el padre/tutor, escaneado o fotografiado."""
    persona = get_object_or_404(Persona, id=persona_id)
    if request.method == "POST" and request.FILES.get("archivo"):
        DocumentoPersona.objects.create(
            persona=persona, tipo="autorizacion_fichaje", archivo=request.FILES["archivo"],
        )
        messages.success(request, "Autorización firmada subida correctamente.")
    return redirect("ficha_persona", persona_id=persona.id)


@login_required
def formulario_08(request, persona_id):
    """Genera el Formulario 08: autorización de fichaje para menores de edad."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    import io

    persona = get_object_or_404(Persona, id=persona_id)
    vinculo = persona.vinculos.filter(fecha_fin__isnull=True).first()
    club = vinculo.club if vinculo else None
    categoria = vinculo.categoria if vinculo else None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=estilos["Title"], alignment=TA_CENTER, fontSize=13)
    estilo_normal = ParagraphStyle("normal", parent=estilos["Normal"], fontSize=10, leading=17, alignment=TA_JUSTIFY)
    elementos = []

    logo_federacion = finders.find("federacion_app/logo.png")
    if logo_federacion:
        elementos.append(Image(logo_federacion, width=2.3 * cm, height=2.3 * cm))
        elementos.append(Spacer(1, 6))

    elementos.append(Paragraph("FORM. 08 — Autorización de fichaje e inclusión en categorías superiores", estilo_titulo))
    elementos.append(Paragraph("(para jugadores de las categorías C7 / C9 / C11 / C13 / C15 / C17 — menores de edad)", estilo_titulo))
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph("Sr. Presidente<br/>Federación Fueguina Fútbol de Salón -Futsal-", estilos["Normal"]))
    elementos.append(Spacer(1, 14))

    texto = (
        "El que suscribe .............................................................................................., "
        "padre-tutor del jugador "
        f"<b>{persona.apellido}, {persona.nombre}</b>, categoría "
        f"<b>{categoria.nombre if categoria else '.....'}</b>, clase "
        f"<b>{persona.fecha_nacimiento.year if persona.fecha_nacimiento else '.....'}</b>, "
        "AUTORIZA al mismo a participar de los torneos de la disciplina FÚTBOL DE SALÓN organizados en forma OFICIAL "
        "por la FEDERACIÓN a nivel Local, Provincial y Nacional para la Entidad Deportiva Afiliada denominada "
        f"<b>{club.nombre if club else '.....'}</b>."
        "<br/><br/>"
        "Esta AUTORIZACIÓN es extensiva para incluir al jugador en UNA / DOS (*) categorías inmediatas superiores, "
        "asumiendo las responsabilidades que me correspondieren."
        "<br/><br/>"
        "A sus efectos, declaro el siguiente domicilio: calle ..................................................... "
        "N°: .............. TE: .................... en esta Ciudad."
        "<br/><br/>"
        "(*) TACHAR LO QUE NO CORRESPONDA"
    )
    elementos.append(Paragraph(texto, estilo_normal))
    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph(f"RÍO GRANDE, ........... de .......................... del {date.today().year}.-", estilos["Normal"]))
    elementos.append(Spacer(1, 30))

    firmas = [
        ["Firma del jugador: _______________________________", f"DNI N°: {persona.documento}"],
        ["", ""],
        ["Firma del padre/tutor: _______________________________", "DNI N°: ________________________"],
        ["", ""],
        ["Domicilio: _______________________________________", "TE/CEL: ____________________________"],
    ]
    tabla_firmas = Table(firmas, colWidths=[11 * cm, 7 * cm])
    tabla_firmas.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 10), ("TOPPADDING", (0, 0), (-1, -1), 8)]))
    elementos.append(tabla_firmas)
    elementos.append(Spacer(1, 30))
    elementos.append(Paragraph("Firma y sello Entidad Deportiva que lo representa: _______________________________", estilos["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="formulario08_{persona.apellido}.pdf"'
    return response


@login_required
def buscar_persona(request):
    """Buscador simple por documento, apellido o número de carnet, para llegar a la ficha."""
    resultados = []
    query = request.GET.get("q", "").strip()
    if query:
        resultados = Persona.objects.filter(
            Q(documento__icontains=query) |
            Q(apellido__icontains=query) |
            Q(numero_carnet__icontains=query)
        )
    return render(request, "federacion_app/buscar_persona.html", {
        "resultados": resultados, "query": query,
    })


# ---------------------------------------------------------------------
# FORMULARIO 12 (inscripción por categoría, jugadores + cuerpo técnico)
# ---------------------------------------------------------------------

@login_required
@user_passes_test(es_delegado)
def formulario_12(request):
    from .forms import TecnicoQuickForm
    from .models import InscripcionTorneo, Torneo

    club = request.user.club
    torneos = Torneo.objects.filter(activo=True)
    categorias = Categoria.objects.all().order_by("nombre")

    torneo_id = request.POST.get("torneo") or request.GET.get("torneo")
    categoria_id = request.POST.get("categoria") or request.GET.get("categoria")
    categoria = get_object_or_404(Categoria, id=categoria_id) if categoria_id else None

    # Alta rápida de cuerpo técnico, sin salir de esta pantalla
    tecnico_form = TecnicoQuickForm()
    if request.method == "POST" and request.POST.get("accion") == "agregar_tecnico":
        tecnico_form = TecnicoQuickForm(request.POST)
        if tecnico_form.is_valid() and categoria:
            documento = tecnico_form.cleaned_data["documento"]
            persona, creada = Persona.objects.get_or_create(
                documento=documento,
                defaults={
                    "tipo": "tecnico",
                    "nombre": tecnico_form.cleaned_data["nombre"],
                    "apellido": tecnico_form.cleaned_data["apellido"],
                    "fecha_nacimiento": date.today(),
                    "rol_tecnico": tecnico_form.cleaned_data["rol_tecnico"],
                },
            )
            if not persona.vinculos.filter(club=club, fecha_fin__isnull=True).exists():
                Vinculo.objects.create(
                    persona=persona, club=club, categoria=categoria, fecha_inicio=date.today()
                )
            messages.success(request, f"{persona} agregado al cuerpo técnico.")
            url = f"/torneos/formulario-12/?torneo={torneo_id or ''}&categoria={categoria_id or ''}"
            return redirect(url)
        else:
            messages.error(request, "Elegí una categoría antes de agregar cuerpo técnico.")

    jugadores, tecnicos = [], []
    ya_cargados_jugadores, ya_cargados_tecnicos = [], []
    torneo_elegido = Torneo.objects.filter(id=torneo_id).first() if torneo_id else None

    if categoria:
        base = Persona.objects.filter(
            vinculos__club=club, vinculos__categoria=categoria, vinculos__fecha_fin__isnull=True
        ).exclude(tipo="jugador", activo=False).distinct()
        jugadores_qs = base.filter(tipo="jugador").order_by("apellido")
        tecnicos_qs = base.filter(tipo="tecnico").order_by("apellido")

        if torneo_elegido:
            # Jugadores que ya tienen inscripción para este torneo (viene de
            # una presentación ya aprobada) no hace falta volver a cargarlos.
            ids_jugadores_cargados = set(
                InscripcionTorneo.objects.filter(club=club, torneo=torneo_elegido).values_list("persona_id", flat=True)
            )
            # Cuerpo técnico ya incluido en alguna presentación aprobada de este torneo.
            from .models import PresentacionFormulario12
            ids_tecnicos_cargados = set(
                PresentacionFormulario12.objects.filter(
                    club=club, torneo=torneo_elegido, estado="aprobado"
                ).values_list("tecnicos__id", flat=True)
            )

            jugadores = jugadores_qs.exclude(id__in=ids_jugadores_cargados)
            ya_cargados_jugadores = jugadores_qs.filter(id__in=ids_jugadores_cargados)
            tecnicos = tecnicos_qs.exclude(id__in=ids_tecnicos_cargados)
            ya_cargados_tecnicos = tecnicos_qs.filter(id__in=ids_tecnicos_cargados)
        else:
            jugadores = jugadores_qs
            tecnicos = tecnicos_qs

    # Generar el/los Formulario 12 en PDF, y dejar la presentación pendiente
    # de que la federación la apruebe (no se inscribe todavía).
    if request.method == "POST" and request.POST.get("accion") == "generar":
        from .models import PresentacionFormulario12

        torneo = get_object_or_404(Torneo, id=torneo_id)
        ids_seleccionados = request.POST.getlist("personas")
        seleccionadas = list(Persona.objects.filter(id__in=ids_seleccionados))
        orden = {int(pid): i for i, pid in enumerate(ids_seleccionados)}
        seleccionadas.sort(key=lambda p: orden.get(p.id, 999))

        if not seleccionadas:
            messages.error(request, "Seleccioná al menos una persona.")
        else:
            presentacion = PresentacionFormulario12.objects.create(
                club=club, torneo=torneo, categoria=categoria, creado_por=request.user,
            )
            presentacion.jugadores.set([p for p in seleccionadas if p.tipo == "jugador"])
            presentacion.tecnicos.set([p for p in seleccionadas if p.tipo == "tecnico"])
            return _generar_pdf_formulario_12(club, torneo, categoria, seleccionadas, request.user)

    return render(request, "federacion_app/formulario_12.html", {
        "torneos": torneos, "categorias": categorias,
        "torneo_id": torneo_id, "categoria": categoria,
        "jugadores": jugadores, "tecnicos": tecnicos,
        "ya_cargados_jugadores": ya_cargados_jugadores, "ya_cargados_tecnicos": ya_cargados_tecnicos,
        "tecnico_form": tecnico_form,
    })


@login_required
@user_passes_test(es_delegado)
def mis_presentaciones_formulario12(request):
    """El delegado ve el estado de cada Formulario 12 presentado y sube el papel firmado."""
    from .models import PresentacionFormulario12
    presentaciones = PresentacionFormulario12.objects.filter(
        club=request.user.club
    ).select_related("torneo", "categoria").prefetch_related("jugadores", "tecnicos")
    return render(request, "federacion_app/mis_presentaciones_formulario12.html", {
        "presentaciones": presentaciones,
    })


@login_required
@user_passes_test(es_delegado)
def subir_formulario12_firmado(request, presentacion_id):
    from .models import PresentacionFormulario12
    presentacion = get_object_or_404(PresentacionFormulario12, id=presentacion_id, club=request.user.club)
    if request.method == "POST" and request.FILES.get("archivo"):
        presentacion.archivo_firmado = request.FILES["archivo"]
        presentacion.save()
        messages.success(request, "Formulario firmado subido. Ahora la federación lo va a revisar.")
    return redirect("mis_presentaciones_formulario12")


def _edad(persona):
    if not persona.fecha_nacimiento:
        return None
    hoy = date.today()
    return hoy.year - persona.fecha_nacimiento.year - (
        (hoy.month, hoy.day) < (persona.fecha_nacimiento.month, persona.fecha_nacimiento.day)
    )


def _menores_sin_autorizacion(presentacion):
    """Jugadores menores de 18 de la presentación que todavía no tienen el Form. 08 firmado subido."""
    faltantes = []
    for p in presentacion.jugadores.all():
        edad = _edad(p)
        if edad is not None and edad < 18:
            if not p.documentos.filter(tipo="autorizacion_fichaje").exists():
                faltantes.append(p)
    return faltantes


@login_required
@user_passes_test(es_federacion)
def presentaciones_formulario12(request):
    """La federación revisa las presentaciones pendientes, filtrando por torneo/club/categoría."""
    from .models import PresentacionFormulario12, Torneo

    torneos = Torneo.objects.filter(activo=True)
    clubes = Club.objects.all().order_by("nombre")
    categorias = Categoria.objects.all().order_by("nombre")

    presentaciones = PresentacionFormulario12.objects.filter(estado="pendiente").select_related(
        "club", "torneo", "categoria"
    ).prefetch_related("jugadores", "tecnicos")

    torneo_id = request.GET.get("torneo")
    club_id = request.GET.get("club")
    categoria_id = request.GET.get("categoria")
    if torneo_id:
        presentaciones = presentaciones.filter(torneo_id=torneo_id)
    if club_id:
        presentaciones = presentaciones.filter(club_id=club_id)
    if categoria_id:
        presentaciones = presentaciones.filter(categoria_id=categoria_id)

    # Para cada presentación, quiénes son menores sin el Form. 08 subido todavía.
    filas = [{"presentacion": p, "menores_sin_autorizacion": _menores_sin_autorizacion(p)} for p in presentaciones]

    return render(request, "federacion_app/presentaciones_formulario12.html", {
        "filas": filas, "torneos": torneos, "clubes": clubes, "categorias": categorias,
        "torneo_id": torneo_id, "club_id": club_id, "categoria_id": categoria_id,
    })


@login_required
@user_passes_test(es_federacion)
def ver_formulario12_presentacion(request, presentacion_id):
    """La federación puede ver el Formulario 12 generado por el delegado, antes de que suban el firmado."""
    from .models import PresentacionFormulario12

    presentacion = get_object_or_404(PresentacionFormulario12, id=presentacion_id)
    personas = list(presentacion.jugadores.all()) + list(presentacion.tecnicos.all())
    return _generar_pdf_formulario_12(
        presentacion.club, presentacion.torneo, presentacion.categoria, personas, request.user
    )


@login_required
@user_passes_test(es_federacion)
def historial_presentaciones_formulario12(request):
    """Historial de presentaciones ya resueltas (aprobadas o rechazadas), con los mismos filtros."""
    from .models import PresentacionFormulario12, Torneo

    torneos = Torneo.objects.filter(activo=True)
    clubes = Club.objects.all().order_by("nombre")
    categorias = Categoria.objects.all().order_by("nombre")

    presentaciones = PresentacionFormulario12.objects.filter(
        estado__in=["aprobado", "rechazado"]
    ).select_related("club", "torneo", "categoria", "aprobado_por").prefetch_related("jugadores", "tecnicos")

    torneo_id = request.GET.get("torneo")
    club_id = request.GET.get("club")
    categoria_id = request.GET.get("categoria")
    if torneo_id:
        presentaciones = presentaciones.filter(torneo_id=torneo_id)
    if club_id:
        presentaciones = presentaciones.filter(club_id=club_id)
    if categoria_id:
        presentaciones = presentaciones.filter(categoria_id=categoria_id)

    presentaciones = presentaciones.order_by("-fecha_aprobacion")

    return render(request, "federacion_app/historial_presentaciones_formulario12.html", {
        "presentaciones": presentaciones, "torneos": torneos, "clubes": clubes, "categorias": categorias,
        "torneo_id": torneo_id, "club_id": club_id, "categoria_id": categoria_id,
    })


@login_required
@user_passes_test(es_federacion)
def resolver_presentacion_formulario12(request, presentacion_id):
    """Aprobar o rechazar una presentación. Al aprobar, recién ahí se inscriben los jugadores al torneo."""
    from .models import PresentacionFormulario12, InscripcionTorneo

    presentacion = get_object_or_404(PresentacionFormulario12, id=presentacion_id)
    if request.method == "POST":
        accion = request.POST.get("accion")
        if accion == "aprobar":
            faltantes = _menores_sin_autorizacion(presentacion)
            if faltantes:
                nombres = ", ".join(str(p) for p in faltantes)
                messages.error(
                    request,
                    f"No se puede aprobar: falta el Formulario 08 firmado de: {nombres}. "
                    f"Pedile al delegado que lo suba desde la ficha de cada jugador."
                )
                return redirect("presentaciones_formulario12")

            for p in presentacion.jugadores.all():
                InscripcionTorneo.objects.get_or_create(
                    persona=p, club=presentacion.club, torneo=presentacion.torneo,
                    defaults={"inscrito_por": presentacion.creado_por},
                )
            presentacion.estado = "aprobado"
            presentacion.aprobado_por = request.user
            presentacion.fecha_aprobacion = date.today()
            presentacion.save()
            messages.success(request, f"Presentación aprobada. {presentacion.jugadores.count()} jugador(es) inscriptos.")
        else:
            presentacion.estado = "rechazado"
            presentacion.motivo_rechazo = request.POST.get("motivo_rechazo", "")
            presentacion.aprobado_por = request.user
            presentacion.fecha_aprobacion = date.today()
            presentacion.save()
            messages.info(request, "Presentación rechazada.")
    return redirect("presentaciones_formulario12")


def _generar_pdf_formulario_12(club, torneo, categoria, personas, usuario):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer, Paragraph, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
    from reportlab.graphics.shapes import Drawing, Rect, String
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=0.4 * cm, bottomMargin=0.4 * cm, leftMargin=0.8 * cm, rightMargin=0.8 * cm,
    )
    estilos = getSampleStyleSheet()
    estilo_chico = ParagraphStyle("chico", parent=estilos["Normal"], fontSize=6.3, alignment=TA_JUSTIFY, leading=7)
    estilo_normal = ParagraphStyle("normal", parent=estilos["Normal"], fontSize=7.5, alignment=TA_LEFT, leading=9)
    estilo_negrita = ParagraphStyle("negrita", parent=estilos["Normal"], fontSize=9, alignment=TA_LEFT, fontName="Helvetica-Bold", leading=10)
    estilo_centro = ParagraphStyle("centro", parent=estilos["Normal"], fontSize=7.5, alignment=TA_CENTER, leading=9)
    elementos = []

    logo_federacion = finders.find("federacion_app/logo.png")
    img_federacion_chico = Image(logo_federacion, width=1.1 * cm, height=1.1 * cm) if logo_federacion else ""
    img_club_chico = Image(club.escudo.path, width=1.1 * cm, height=1.1 * cm) if club.escudo else ""

    texto_legal = (
        "Los integrantes de la presente planilla declaran conocer y aceptar las condiciones absolutas de "
        "amateurismo del Futsal / Fútbol de Salón, conocer y aceptar las reglas de juego y el alto grado de "
        "competitividad alcanzado en los eventos organizados por las entidades afiliadas a la Confederación "
        "Argentina de Futsal, conocer y aceptar la existencia de potenciales riesgos para la salud en la práctica "
        "de nuestro deporte, declaran conocer y aceptar las obligaciones Estatutarias y Reglamentarias a las que "
        "adhiere el jugador FEDERADO de la C.A.F.S., también se comprometen a mantener vigente la cobertura del "
        "seguro obligatorio y que por ello liberan de toda responsabilidad a la Entidad, dirigentes y miembros de "
        "los diferentes cuerpos que la integran, por cualquier suceso que pudiera ocurrir en los traslados, actos, "
        "durante y después de la competencia a la que corresponde esta planilla de fichaje."
    )

    # Solo los jugadores ocupan las 20 líneas numeradas; el cuerpo técnico
    # va aparte, en 4 renglones fijos (uno por rol).
    jugadores = [p for p in personas if p.tipo == "jugador"]
    tecnicos = [p for p in personas if p.tipo == "tecnico"]
    tecnico_por_rol = {}
    for t in tecnicos:
        if t.rol_tecnico and t.rol_tecnico not in tecnico_por_rol:
            tecnico_por_rol[t.rol_tecnico] = t

    grupos = [jugadores[i:i + 20] for i in range(0, len(jugadores), 20)] or [[]]

    for numero_form, grupo in enumerate(grupos, start=1):
        # --- Encabezado ---
        encabezado_izq = Paragraph(
            "<b>FEDERACIÓN FUEGUINA DE FUTBOL DE SALON -FUTSAL-</b><br/>"
            "Afiliada a la CONFEDERACION ARGENTINA DE FUTBOL DE SALON",
            estilo_normal,
        )
        encabezado_centro = Table([["FORM-12"]], colWidths=[3 * cm], rowHeights=[1 * cm])
        encabezado_centro.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
        ]))
        encabezado_der = Paragraph(
            f"Fecha presentación: {date.today().strftime('%d/%m/%Y')}<br/>"
            f"<b>Planilla Fichaje</b><br/>"
            f"Denominación real/legal Club/ Agrupación<br/>"
            f"<b>{club.nombre.upper()}</b>",
            estilo_normal,
        )
        tabla_encabezado = Table(
            [[encabezado_izq, encabezado_centro, encabezado_der]],
            colWidths=[9 * cm, 3.5 * cm, 12 * cm],
        )
        tabla_encabezado.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        elementos.append(tabla_encabezado)
        elementos.append(Spacer(1, 3))

        tabla_datos = Table(
            [[f"Provincia: Tierra del Fuego   Ciudad: Río Grande",
              f"EVENTO DEPORTIVO: {torneo}",
              f"CATEGORIA: {categoria}"]],
            colWidths=[9 * cm, 8 * cm, 7.5 * cm],
        )
        tabla_datos.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.7, colors.black),
            ("INNERGRID", (0, 0), (-1, -1), 0.7, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elementos.append(tabla_datos)
        elementos.append(Spacer(1, 3))
        elementos.append(Paragraph(texto_legal, estilo_chico))
        elementos.append(Spacer(1, 3))

        # --- Tabla principal ---
        encabezados_tabla = ["ORDEN", "CARGO", "CARNET", "APELLIDOS", "NOMBRES",
                              "DOCUMENTO (DNI o CEDULA)", "FECHA NACIMIENTO", "Otros / Varios"]
        filas = [encabezados_tabla]

        for i in range(1, 21):
            if i <= len(grupo):
                p = grupo[i - 1]
                filas.append([
                    str(i), "JUGADOR", p.numero_carnet or "", p.apellido, p.nombre,
                    p.documento, p.fecha_nacimiento.strftime("%d/%m/%Y") if p.fecha_nacimiento else "", "",
                ])
            else:
                filas.append([str(i), "JUGADOR", "", "", "", "", "", ""])

        # Los 4 renglones fijos de cuerpo técnico, sin número de orden
        for rol_key, etiqueta in [("dt", "Director Tecnico"), ("ayudante", "Ayu. Tecnico"),
                                    ("pf", "Preparador Fisico"), ("delegado", "DELEGADO")]:
            t = tecnico_por_rol.get(rol_key)
            if t:
                filas.append(["", etiqueta, t.numero_carnet or "", t.apellido, t.nombre,
                               t.documento, t.fecha_nacimiento.strftime("%d/%m/%Y") if t.fecha_nacimiento else "", ""])
            else:
                filas.append(["", etiqueta, "", "", "", "", "", ""])

        tabla = Table(
            filas,
            colWidths=[1.3 * cm, 2.7 * cm, 2 * cm, 4 * cm, 4 * cm, 3.8 * cm, 3 * cm, 2.7 * cm],
            repeatRows=1,
        )
        tabla.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f2a4a")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (0, 0), (2, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 0.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0.5),
        ]))
        elementos.append(tabla)
        elementos.append(Spacer(1, 3))

        # --- Certificación, sellos y firmas ---
        cert_izq = Paragraph(
            "CERTIFICO QUE LOS DATOS ELEVADOS SON VERDADEROS<br/>Y AVALADOS POR NUESTRA ENTIDAD",
            estilo_centro,
        )
        cert_der = Paragraph(
            "LOS DATOS ELEVADOS SON CONSIDERADOS VERDADEROS<br/>Y AVALADOS POR NUESTRA ENTIDAD",
            estilo_centro,
        )
        firma_izq = Paragraph("_______________________________<br/>Firma Presidente Club", estilo_centro)
        firma_der = Paragraph("_______________________________<br/>Firma Responsable Fo.Fu.Futsal", estilo_centro)
        sello_club = Paragraph("SELLO CLUB", estilo_centro)

        # Sello de la federación, dibujado como una estampilla real (no texto plano):
        # arriba el nombre de la federación, en el medio la fecha, abajo "RECIBIDO".
        meses_abrev = {1: "ENE", 2: "FEB", 3: "MAR", 4: "ABR", 5: "MAY", 6: "JUN",
                       7: "JUL", 8: "AGO", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DIC"}
        hoy = date.today()
        fecha_sello = f"{hoy.day:02d} {meses_abrev[hoy.month]} {hoy.year}"

        ancho_sello, alto_sello = 56, 40
        sello_federacion = Drawing(ancho_sello, alto_sello)
        sello_federacion.add(Rect(
            1, 1, ancho_sello - 2, alto_sello - 2, rx=6, ry=6,
            strokeColor=colors.black, strokeWidth=1, fillColor=None,
        ))
        sello_federacion.add(String(ancho_sello / 2, alto_sello - 9, "FEDERACIÓN",
                                     fontName="Helvetica-Bold", fontSize=5, textAnchor="middle", fillColor=colors.black))
        sello_federacion.add(String(ancho_sello / 2, alto_sello - 15, "FUEGUINA",
                                     fontName="Helvetica-Bold", fontSize=5, textAnchor="middle", fillColor=colors.black))
        sello_federacion.add(String(ancho_sello / 2, alto_sello - 21, "FÚTBOL DE SALÓN",
                                     fontName="Helvetica-Bold", fontSize=4.6, textAnchor="middle", fillColor=colors.black))
        sello_federacion.add(Rect(5, alto_sello - 30, ancho_sello - 10, 9,
                                   strokeColor=colors.black, strokeWidth=0.7, fillColor=None))
        sello_federacion.add(String(ancho_sello / 2, alto_sello - 27.5, fecha_sello,
                                     fontName="Helvetica-Bold", fontSize=5.4, textAnchor="middle", fillColor=colors.red))
        sello_federacion.add(String(ancho_sello / 2, 4, "RECIBIDO",
                                     fontName="Helvetica-Bold", fontSize=5.8, textAnchor="middle", fillColor=colors.black))

        tabla_firmas = Table(
            [
                [cert_izq, img_club_chico, img_federacion_chico, cert_der],
                ["", "", "", ""],  # espacio en blanco para firmar a mano
                [firma_izq, sello_club, "", [sello_federacion, firma_der]],
            ],
            colWidths=[8 * cm, 2.5 * cm, 2.5 * cm, 8 * cm],
            rowHeights=[None, 0.9 * cm, None],
        )
        tabla_firmas.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]))
        elementos.append(tabla_firmas)

        if numero_form < len(grupos):
            elementos.append(PageBreak())

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="formulario12_{club.nombre}_{categoria}.pdf"'
    return response


@login_required
@user_passes_test(es_delegado)
def formulario_07(request):
    """Pantalla para cargar hasta 4 delegados deportivos y generar el Formulario 07."""
    if request.method == "POST":
        return _generar_pdf_formulario_07(request, request.user.club)
    return render(request, "federacion_app/formulario_07.html", {"club": request.user.club})


def _generar_pdf_formulario_07(request, club):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Image, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from django.contrib.staticfiles import finders
    from django.http import HttpResponse
    import io

    temporada = request.POST.get("temporada", "")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.3 * cm, bottomMargin=1.3 * cm)
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=estilos["Title"], alignment=TA_CENTER, fontSize=13)
    estilo_normal = ParagraphStyle("normal", parent=estilos["Normal"], fontSize=9, leading=13, alignment=TA_JUSTIFY)
    estilo_delegado = ParagraphStyle("delegado", parent=estilos["Normal"], fontSize=9.5, leading=15)
    elementos = []

    logo_federacion = finders.find("federacion_app/logo.png")
    if logo_federacion:
        elementos.append(Image(logo_federacion, width=2 * cm, height=2 * cm))
        elementos.append(Spacer(1, 6))

    elementos.append(Paragraph("FORM. 07 — Designación Delegados Deportivos", estilo_titulo))
    elementos.append(Spacer(1, 8))
    elementos.append(Paragraph("Sr. Presidente<br/>Federación Fueguina Fútbol de Salón -Futsal-", estilos["Normal"]))
    elementos.append(Spacer(1, 12))

    texto_intro = (
        f"Los que suscriben, integrantes de la ENTIDAD deportiva denominada <b>{club.nombre}</b>, tenemos el agrado "
        "de dirigirnos a Ud. con el objeto de informarle la designación de los DELEGADOS DEPORTIVOS que nos "
        f"representarán en las reuniones ordinarias que la competencia local, provincial de la temporada "
        f"<b>{temporada}</b>, siendo sus datos filiatorios los abajo detallados y delegando en los mismos la mayor "
        "facultad y responsabilidad ante la Federación Fueguina Fútbol de Salón -Futsal-, anticipando que los "
        "mismos declaran conocer y aceptar las siguientes condiciones, tal lo hiciera nuestra ENTIDAD, a saber:"
    )
    elementos.append(Paragraph(texto_intro, estilo_normal))
    elementos.append(Spacer(1, 8))

    condiciones = [
        "Reconocemos y aceptamos que el Fútbol de Salón o Futsal es patrimonio exclusivo en el país de la "
        "Confederación Argentina de Fútbol de Salón / Futsal.",
        "Declaramos conocer y aceptar el Reglamento de Juego y las condiciones exigidas por la CAFS en sus "
        "normas de procedimiento.",
        "Declaramos conocer y aceptar las normas vigentes para asociarnos a la Federación Fueguina Fútbol de "
        "Salón -Futsal-.",
        "Declaramos conocer y aceptar el alto grado de competitividad alcanzado en los torneos organizados por "
        "la entidad y la CAFS.",
        "Declaramos conocer y aceptar la existencia de potenciales riesgos para la salud en la práctica de "
        "Fútbol de Salón o Futsal.",
        "Declaramos conocer y aceptar las obligaciones Estatutarias y Reglamentarias de la Federación y la CAFS.",
        "Nos comprometemos a mantener vigente la cobertura de seguros contra accidentes personales exigida.",
        "Liberamos de toda responsabilidad a la CAFS, sus afiliadas y dirigentes por cualquier accidente que "
        "pudiera ocurrir durante traslados, antes, durante y después de cualquier competencia.",
    ]
    for c in condiciones:
        elementos.append(Paragraph(f"• {c}", estilo_normal))
        elementos.append(Spacer(1, 3))

    elementos.append(Spacer(1, 10))

    etiquetas = ["1er DELEGADO DEPORTIVO", "2do DELEGADO DEPORTIVO", "3er DELEGADO DEPORTIVO", "4to DELEGADO DEPORTIVO"]
    for i, etiqueta in enumerate(etiquetas, start=1):
        nombre = request.POST.get(f"nombre_{i}", "")
        dni = request.POST.get(f"dni_{i}", "")
        fecha_nac = request.POST.get(f"fecha_nac_{i}", "")
        domicilio = request.POST.get(f"domicilio_{i}", "")
        localidad = request.POST.get(f"localidad_{i}", "")
        telefono = request.POST.get(f"telefono_{i}", "")

        if not nombre:
            continue

        elementos.append(Paragraph(f"<b>{etiqueta}</b>", estilo_delegado))
        elementos.append(Paragraph(
            f"Sr./a: {nombre} — DNI N°: {dni} — Fe. Nac.: {fecha_nac}<br/>"
            f"Domicilio: {domicilio} — Localidad: {localidad} — TE fijo o CEL: {telefono}",
            estilo_delegado,
        ))
        elementos.append(Spacer(1, 8))

    elementos.append(Spacer(1, 20))
    elementos.append(Paragraph(f"Firma y sello entidad Deportiva: {club.nombre} — ___________________________________", estilos["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="formulario07_{club.nombre}.pdf"'
    return response


# ---------------------------------------------------------------------
# TARJETAS Y SANCIONES
# ---------------------------------------------------------------------

# Montos por umbral, según el reglamento de la federación. Si se supera
# la 4ta ocurrencia, se sigue usando el último monto de la tabla.
MONTOS_AMARILLA = [20000, 25000, 30000, 35000]
MONTOS_AZUL_INDIRECTA = [30000, 35000, 40000, 45000]
MONTOS_AZUL_DIRECTA = [40000, 45000, 50000, 55000]


def _monto_por_ocurrencia(lista_montos, ocurrencia):
    indice = min(ocurrencia - 1, len(lista_montos) - 1)
    return lista_montos[indice]


def _revisar_umbrales_sancion(persona, club):
    """
    Después de cargar una tarjeta, revisa si el jugador llegó a algún
    umbral nuevo (4ta/8va/... amarilla, 2da/4ta/... azul indirecta, o
    cualquier azul directa) y genera la sanción correspondiente si
    todavía no existía.
    """
    from .models import Tarjeta, SancionTarjeta

    # Amarillas: cada grupo de 4
    total_amarillas = Tarjeta.objects.filter(persona=persona, tipo="amarilla").count()
    if total_amarillas and total_amarillas % 4 == 0:
        ocurrencia = total_amarillas // 4
        SancionTarjeta.objects.get_or_create(
            persona=persona, tipo_tarjeta="amarilla", numero_ocurrencia=ocurrencia,
            defaults={"club": club, "monto": _monto_por_ocurrencia(MONTOS_AMARILLA, ocurrencia)},
        )

    # Azules indirectas: cada grupo de 2
    total_indirectas = Tarjeta.objects.filter(persona=persona, tipo="azul_indirecta").count()
    if total_indirectas and total_indirectas % 2 == 0:
        ocurrencia = total_indirectas // 2
        SancionTarjeta.objects.get_or_create(
            persona=persona, tipo_tarjeta="azul_indirecta", numero_ocurrencia=ocurrencia,
            defaults={"club": club, "monto": _monto_por_ocurrencia(MONTOS_AZUL_INDIRECTA, ocurrencia)},
        )

    # Azules directas: cada tarjeta individual ya sanciona
    total_directas = Tarjeta.objects.filter(persona=persona, tipo="azul_directa").count()
    if total_directas:
        SancionTarjeta.objects.get_or_create(
            persona=persona, tipo_tarjeta="azul_directa", numero_ocurrencia=total_directas,
            defaults={"club": club, "monto": _monto_por_ocurrencia(MONTOS_AZUL_DIRECTA, total_directas)},
        )


@login_required
@user_passes_test(es_federacion)
def cargar_tarjeta(request):
    """La federación carga una tarjeta a un jugador después de un partido."""
    from .models import Tarjeta, Torneo

    clubes = Club.objects.all().order_by("nombre")
    torneos = Torneo.objects.filter(activo=True)
    club_id = request.POST.get("club") or request.GET.get("club")
    club_elegido = Club.objects.filter(id=club_id).first() if club_id else None

    jugadores = []
    if club_elegido:
        jugadores = Persona.objects.filter(
            vinculos__club=club_elegido, vinculos__fecha_fin__isnull=True, tipo="jugador"
        ).distinct().order_by("apellido")

    if request.method == "POST" and request.POST.get("accion") == "cargar":
        persona = get_object_or_404(Persona, id=request.POST.get("persona"))
        tipo = request.POST.get("tipo")
        fecha_partido = request.POST.get("fecha_partido")
        torneo_id = request.POST.get("torneo")

        Tarjeta.objects.create(
            persona=persona, club=club_elegido, tipo=tipo, fecha_partido=fecha_partido,
            torneo_id=torneo_id or None, observacion=request.POST.get("observacion", ""),
            cargada_por=request.user,
        )
        _revisar_umbrales_sancion(persona, club_elegido)
        messages.success(request, f"Tarjeta cargada para {persona}.")
        return redirect(f"/tarjetas/cargar/?club={club_elegido.id}")

    return render(request, "federacion_app/cargar_tarjeta.html", {
        "clubes": clubes, "torneos": torneos, "club_elegido": club_elegido, "jugadores": jugadores,
    })


@login_required
@user_passes_test(es_federacion)
def sanciones_pendientes(request):
    """La federación ve las sanciones pendientes y las resuelve (pagado / cumplida)."""
    from .models import SancionTarjeta
    sanciones = SancionTarjeta.objects.filter(estado="pendiente").select_related("persona", "club")
    return render(request, "federacion_app/sanciones_pendientes.html", {"sanciones": sanciones})


@login_required
@user_passes_test(es_federacion)
def resolver_sancion(request, sancion_id):
    from .models import SancionTarjeta
    sancion = get_object_or_404(SancionTarjeta, id=sancion_id)
    if request.method == "POST":
        accion = request.POST.get("accion")
        if accion in ("pagado", "cumplido"):
            sancion.estado = accion
            sancion.fecha_resolucion = date.today()
            sancion.resuelto_por = request.user
            sancion.save()
            messages.success(request, f"Sanción de {sancion.persona} marcada como {sancion.get_estado_display()}.")
    return redirect("sanciones_pendientes")


@login_required
@user_passes_test(es_delegado)
def mis_sanciones(request):
    """El delegado ve las sanciones (por tarjetas y disciplinarias) de su club y sube el comprobante de pago."""
    from .models import SancionTarjeta, SancionDisciplinaria
    sanciones = SancionTarjeta.objects.filter(club=request.user.club).select_related("persona")
    sanciones_disciplinarias = SancionDisciplinaria.objects.filter(club=request.user.club).select_related("persona")
    return render(request, "federacion_app/mis_sanciones.html", {
        "sanciones": sanciones, "sanciones_disciplinarias": sanciones_disciplinarias,
    })


@login_required
@user_passes_test(es_delegado)
def mis_tarjetas(request):
    """El delegado ve todas las tarjetas de su club, con un resumen de cuánto le falta a cada jugador para el próximo umbral."""
    from .models import Tarjeta
    from collections import defaultdict

    tarjetas = Tarjeta.objects.filter(club=request.user.club).select_related("persona", "torneo").order_by("-fecha_partido")

    # Resumen por jugador: cuántas tiene de cada tipo, y cuántas le faltan
    # para el próximo umbral (4 amarillas, 2 azules indirectas).
    conteo = defaultdict(lambda: {"persona": None, "amarilla": 0, "azul_indirecta": 0, "azul_directa": 0})
    for t in tarjetas:
        conteo[t.persona_id]["persona"] = t.persona
        conteo[t.persona_id][t.tipo] += 1

    resumen = []
    for datos in conteo.values():
        resumen.append({
            "persona": datos["persona"],
            "amarillas": datos["amarilla"],
            "faltan_amarilla": (4 - datos["amarilla"] % 4) % 4 or 4,
            "azules_indirectas": datos["azul_indirecta"],
            "faltan_indirecta": (2 - datos["azul_indirecta"] % 2) % 2 or 2,
            "azules_directas": datos["azul_directa"],
        })
    resumen.sort(key=lambda r: str(r["persona"]))

    return render(request, "federacion_app/mis_tarjetas.html", {"tarjetas": tarjetas, "resumen": resumen})


@login_required
@user_passes_test(es_delegado)
def subir_comprobante_sancion(request, sancion_id):
    from .models import SancionTarjeta
    sancion = get_object_or_404(SancionTarjeta, id=sancion_id, club=request.user.club)
    if request.method == "POST" and request.FILES.get("archivo"):
        sancion.comprobante_pago = request.FILES["archivo"]
        sancion.save()
        messages.success(request, "Comprobante subido. La federación lo va a revisar antes del próximo partido.")
    return redirect("mis_sanciones")


# ---------------------------------------------------------------------
# CONSEJO DE DISCIPLINA
# ---------------------------------------------------------------------

def _es_consejo_o_federacion(usuario):
    return usuario.is_authenticated and usuario.rol in ("consejo_disciplina", "federacion")


@login_required
@user_passes_test(_es_consejo_o_federacion)
def panel_disciplina(request):
    """
    Único módulo al que accede el Consejo de Disciplina: buscar un
    jugador (por documento o apellido, sin necesitar el buscador
    general) y cargarle una sanción con su informe adjunto.
    """
    from .models import SancionDisciplinaria

    query = request.GET.get("q", "").strip()
    resultados = []
    if query:
        resultados = Persona.objects.filter(
            Q(documento__icontains=query) | Q(apellido__icontains=query)
        )[:15]

    persona_id = request.GET.get("persona")
    persona_elegida = Persona.objects.filter(id=persona_id).first() if persona_id else None

    if request.method == "POST":
        from .models import Notificacion

        persona = get_object_or_404(Persona, id=request.POST.get("persona"))
        vinculo = persona.vinculos.filter(fecha_fin__isnull=True).first()
        club = vinculo.club if vinculo else None

        sancion = SancionDisciplinaria.objects.create(
            persona=persona,
            club=club,
            motivo=request.POST.get("motivo", ""),
            informe=request.FILES.get("informe"),
            fecha_sancion=request.POST.get("fecha_sancion"),
            cantidad_fechas=request.POST.get("cantidad_fechas") or None,
            cantidad_anios=request.POST.get("cantidad_anios") or None,
            cargada_por=request.user,
        )

        # Si se tildó, se manda además una notificación formal al club con
        # la resolución/notificación adjunta (puede ser el mismo archivo
        # del informe, o uno distinto si se subió aparte).
        if request.POST.get("notificar_club") and club:
            archivo_notificacion = request.FILES.get("archivo_notificacion") or request.FILES.get("informe")
            if archivo_notificacion and hasattr(archivo_notificacion, "seek"):
                archivo_notificacion.seek(0)
            notificacion = Notificacion.objects.create(
                titulo=request.POST.get("titulo_notificacion") or f"Resolución disciplinaria — {persona}",
                mensaje=request.POST.get("mensaje_notificacion") or sancion.motivo,
                archivo=archivo_notificacion,
                creada_por=request.user,
            )
            notificacion.destinatarios.set([club])

        messages.success(request, f"Sanción disciplinaria cargada para {persona}.")
        return redirect("panel_disciplina")

    sanciones = SancionDisciplinaria.objects.select_related("persona", "club").order_by("-fecha_sancion")[:50]

    return render(request, "federacion_app/panel_disciplina.html", {
        "query": query, "resultados": resultados, "persona_elegida": persona_elegida,
        "sanciones": sanciones,
    })


@login_required
@user_passes_test(_es_consejo_o_federacion)
def resolver_sancion_disciplinaria(request, sancion_id):
    from .models import SancionDisciplinaria
    sancion = get_object_or_404(SancionDisciplinaria, id=sancion_id)
    if request.method == "POST":
        sancion.estado = "cumplida"
        sancion.save()
        messages.success(request, f"Sanción de {sancion.persona} marcada como cumplida.")
    return redirect("panel_disciplina")


# ---------------------------------------------------------------------
# NOTIFICACIONES
# ---------------------------------------------------------------------

@login_required
@user_passes_test(es_federacion)
def crear_notificacion(request):
    """La federación redacta un aviso y elige a qué clubes va dirigido (o a todos)."""
    from .models import Notificacion, NotificacionAcuse

    clubes = Club.objects.all().order_by("nombre")

    if request.method == "POST":
        titulo = request.POST.get("titulo", "").strip()
        mensaje = request.POST.get("mensaje", "").strip()
        club_ids = request.POST.getlist("clubes")

        if not titulo or not mensaje:
            messages.error(request, "Completá título y mensaje.")
        else:
            notificacion = Notificacion.objects.create(
                titulo=titulo, mensaje=mensaje, archivo=request.FILES.get("archivo"), creada_por=request.user,
            )
            if club_ids:
                notificacion.destinatarios.set(club_ids)
            messages.success(request, "Notificación enviada.")
            return redirect("crear_notificacion")

    # Al entrar a esta pantalla, se marcan como vistos todos los acuses
    # (así el badge del menú se limpia).
    NotificacionAcuse.objects.filter(
        fecha_acuse__isnull=False, visto_por_federacion=False
    ).update(visto_por_federacion=True)

    enviadas = Notificacion.objects.prefetch_related(
        "destinatarios", "acuses__club", "acuses__acusado_por"
    ).order_by("-fecha_creacion")[:30]

    enviadas_con_estado = []
    for n in enviadas:
        clubes_destino = list(n.destinatarios.all()) or list(clubes)
        acuses_por_club = {a.club_id: a for a in n.acuses.all()}
        filas_clubes = []
        for c in clubes_destino:
            acuse = acuses_por_club.get(c.id)
            filas_clubes.append({"club": c, "acuse": acuse})
        enviadas_con_estado.append({"notificacion": n, "filas_clubes": filas_clubes})

    return render(request, "federacion_app/crear_notificacion.html", {
        "clubes": clubes, "enviadas": enviadas_con_estado,
    })


@login_required
@user_passes_test(es_delegado)
def notificaciones(request):
    """El delegado ve los avisos dirigidos a su club (o a todos), y se marcan como leídos al entrar."""
    from .models import Notificacion, NotificacionAcuse

    club = request.user.club
    todas = Notificacion.objects.prefetch_related("destinatarios").order_by("-fecha_creacion")
    propias = [n for n in todas if n.es_para(club)]

    # Al entrar a esta pantalla, se marcan todas como leídas por este usuario.
    for n in propias:
        n.leida_por.add(request.user)

    # Trae (o prepara vacío) el acuse de cada notificación para este club.
    filas = []
    for n in propias:
        acuse, _ = NotificacionAcuse.objects.get_or_create(notificacion=n, club=club)
        filas.append({"notificacion": n, "acuse": acuse})

    return render(request, "federacion_app/notificaciones.html", {"filas": filas})


@login_required
@user_passes_test(es_delegado)
def acusar_notificacion(request, notificacion_id):
    """El delegado acusa recibo de una notificación y, opcionalmente, responde."""
    from .models import Notificacion, NotificacionAcuse

    notificacion = get_object_or_404(Notificacion, id=notificacion_id)
    acuse, _ = NotificacionAcuse.objects.get_or_create(notificacion=notificacion, club=request.user.club)

    if request.method == "POST":
        if not acuse.fecha_acuse:
            acuse.acusado_por = request.user
            acuse.fecha_acuse = date.today()

        respuesta = request.POST.get("respuesta", "").strip()
        if respuesta:
            acuse.respuesta = respuesta
            acuse.fecha_respuesta = date.today()

        acuse.save()
        messages.success(request, "Acuse de recibo registrado.")

    return redirect("notificaciones")


# ---------------------------------------------------------------------
# INSTITUCIONAL
# ---------------------------------------------------------------------

@login_required
def institucional(request):
    """
    Reglamentos, formularios y circulares. Cualquier usuario logueado
    puede ver y descargar; solo la federación puede subir o borrar.
    """
    from .models import DocumentoInstitucional

    if request.method == "POST" and es_federacion(request.user):
        accion = request.POST.get("accion")
        if accion == "subir" and request.FILES.get("archivo"):
            DocumentoInstitucional.objects.create(
                titulo=request.POST.get("titulo", "").strip() or request.FILES["archivo"].name,
                tipo=request.POST.get("tipo", "otro"),
                archivo=request.FILES["archivo"],
                descripcion=request.POST.get("descripcion", ""),
                subido_por=request.user,
            )
            messages.success(request, "Documento subido.")
        elif accion == "borrar":
            DocumentoInstitucional.objects.filter(id=request.POST.get("documento_id")).delete()
            messages.info(request, "Documento eliminado.")
        return redirect("institucional")

    documentos = DocumentoInstitucional.objects.all()
    return render(request, "federacion_app/institucional.html", {"documentos": documentos})


# ---------------------------------------------------------------------
# GOLEADORES
# ---------------------------------------------------------------------

@login_required
@user_passes_test(es_federacion)
def cargar_gol(request):
    """La federación carga los goles de un jugador en un partido."""
    from .models import Gol, Torneo

    clubes = Club.objects.all().order_by("nombre")
    torneos = Torneo.objects.filter(activo=True)
    club_id = request.POST.get("club") or request.GET.get("club")
    club_elegido = Club.objects.filter(id=club_id).first() if club_id else None

    jugadores = []
    if club_elegido:
        jugadores = Persona.objects.filter(
            vinculos__club=club_elegido, vinculos__fecha_fin__isnull=True, tipo="jugador"
        ).distinct().order_by("apellido")

    if request.method == "POST" and request.POST.get("accion") == "cargar":
        persona = get_object_or_404(Persona, id=request.POST.get("persona"))
        Gol.objects.create(
            persona=persona, club=club_elegido,
            torneo_id=request.POST.get("torneo") or None,
            fecha_partido=request.POST.get("fecha_partido"),
            cantidad=request.POST.get("cantidad") or 1,
            cargado_por=request.user,
        )
        messages.success(request, f"Gol(es) cargado(s) para {persona}.")
        return redirect(f"/goleadores/cargar/?club={club_elegido.id}")

    return render(request, "federacion_app/cargar_gol.html", {
        "clubes": clubes, "torneos": torneos, "club_elegido": club_elegido, "jugadores": jugadores,
    })


@login_required
def goleadores(request):
    """Ranking de goleadores, filtrable por torneo y categoría. Ven todos los usuarios logueados."""
    from collections import defaultdict
    from .models import Gol, Torneo

    torneos = Torneo.objects.all().order_by("-temporada")
    categorias = Categoria.objects.all().order_by("nombre")

    torneo_id = request.GET.get("torneo")
    categoria_id = request.GET.get("categoria")

    goles = Gol.objects.select_related("persona", "club", "torneo")
    if torneo_id:
        goles = goles.filter(torneo_id=torneo_id)

    acumulado = defaultdict(lambda: {"persona": None, "club": None, "total": 0})
    for g in goles:
        if categoria_id:
            vinculo = g.persona.vinculos.filter(club=g.club, fecha_fin__isnull=True).first()
            if not vinculo or str(vinculo.categoria_id) != categoria_id:
                continue
        acumulado[g.persona_id]["persona"] = g.persona
        acumulado[g.persona_id]["club"] = g.club
        acumulado[g.persona_id]["total"] += g.cantidad

    ranking = sorted(acumulado.values(), key=lambda r: r["total"], reverse=True)

    return render(request, "federacion_app/goleadores.html", {
        "ranking": ranking, "torneos": torneos, "categorias": categorias,
        "torneo_id": torneo_id, "categoria_id": categoria_id,
    })


# ---------------------------------------------------------------------
# VALLA MENOS VENCIDA
# ---------------------------------------------------------------------

@login_required
@user_passes_test(es_federacion)
def cargar_gol_recibido(request):
    """La federación carga los goles recibidos por un jugador (arquero) en un partido."""
    from .models import GolRecibido, Torneo

    clubes = Club.objects.all().order_by("nombre")
    torneos = Torneo.objects.filter(activo=True)
    club_id = request.POST.get("club") or request.GET.get("club")
    club_elegido = Club.objects.filter(id=club_id).first() if club_id else None

    jugadores = []
    if club_elegido:
        jugadores = Persona.objects.filter(
            vinculos__club=club_elegido, vinculos__fecha_fin__isnull=True, tipo="jugador"
        ).distinct().order_by("apellido")

    if request.method == "POST" and request.POST.get("accion") == "cargar":
        persona = get_object_or_404(Persona, id=request.POST.get("persona"))
        GolRecibido.objects.create(
            persona=persona, club=club_elegido,
            torneo_id=request.POST.get("torneo") or None,
            fecha_partido=request.POST.get("fecha_partido"),
            cantidad=request.POST.get("cantidad") or 1,
            cargado_por=request.user,
        )
        messages.success(request, f"Gol(es) recibido(s) cargado(s) para {persona}.")
        return redirect(f"/valla-menos-vencida/cargar/?club={club_elegido.id}")

    return render(request, "federacion_app/cargar_gol_recibido.html", {
        "clubes": clubes, "torneos": torneos, "club_elegido": club_elegido, "jugadores": jugadores,
    })


@login_required
def valla_menos_vencida(request):
    """Ranking de valla menos vencida (menos goles recibidos), filtrable por torneo y categoría."""
    from collections import defaultdict
    from .models import GolRecibido, Torneo

    torneos = Torneo.objects.all().order_by("-temporada")
    categorias = Categoria.objects.all().order_by("nombre")

    torneo_id = request.GET.get("torneo")
    categoria_id = request.GET.get("categoria")

    goles = GolRecibido.objects.select_related("persona", "club", "torneo")
    if torneo_id:
        goles = goles.filter(torneo_id=torneo_id)

    acumulado = defaultdict(lambda: {"persona": None, "club": None, "total": 0})
    for g in goles:
        if categoria_id:
            vinculo = g.persona.vinculos.filter(club=g.club, fecha_fin__isnull=True).first()
            if not vinculo or str(vinculo.categoria_id) != categoria_id:
                continue
        acumulado[g.persona_id]["persona"] = g.persona
        acumulado[g.persona_id]["club"] = g.club
        acumulado[g.persona_id]["total"] += g.cantidad

    # Menos goles recibidos primero (a diferencia del ranking de goleadores).
    ranking = sorted(acumulado.values(), key=lambda r: r["total"])

    return render(request, "federacion_app/valla_menos_vencida.html", {
        "ranking": ranking, "torneos": torneos, "categorias": categorias,
        "torneo_id": torneo_id, "categoria_id": categoria_id,
    })


# ---------------------------------------------------------------------
# PUNITORIOS
# ---------------------------------------------------------------------

@login_required
@user_passes_test(es_federacion)
def cargar_punitorio(request):
    """La federación aplica un punitorio a un club, con monto y motivo manuales, y notifica al club."""
    from .models import ConceptoPunitorio, Punitorio, Notificacion

    clubes = Club.objects.all().order_by("nombre")
    conceptos = ConceptoPunitorio.objects.filter(activo=True)

    if request.method == "POST":
        club_elegido = get_object_or_404(Club, id=request.POST.get("club"))
        concepto_id = request.POST.get("concepto") or None
        motivo = request.POST.get("motivo", "").strip()
        monto = request.POST.get("monto")

        if not motivo or not monto:
            messages.error(request, "Completá el motivo y el monto.")
        else:
            punitorio = Punitorio.objects.create(
                club=club_elegido, concepto_id=concepto_id, motivo=motivo, monto=monto,
                cargado_por=request.user,
            )

            notificacion = Notificacion.objects.create(
                titulo=f"Punitorio aplicado — ${monto}",
                mensaje=f"{motivo}\n\nPara subir el comprobante de pago, andá al menú → \"Punitorios\".",
                creada_por=request.user,
            )
            notificacion.destinatarios.set([club_elegido])

            messages.success(request, f"Punitorio de ${monto} aplicado a {club_elegido.nombre} y notificado.")
            return redirect("cargar_punitorio")

    return render(request, "federacion_app/cargar_punitorio.html", {
        "clubes": clubes, "conceptos": conceptos,
    })


@login_required
@user_passes_test(es_federacion)
def punitorios_pendientes(request):
    """La federación ve los punitorios pendientes de cobro y los marca como pagados."""
    from .models import Punitorio
    punitorios = Punitorio.objects.filter(estado="pendiente").select_related("club", "concepto")
    return render(request, "federacion_app/punitorios_pendientes.html", {"punitorios": punitorios})


@login_required
@user_passes_test(es_federacion)
def historial_punitorios(request):
    """Historial de punitorios ya pagados, filtrable por club."""
    from .models import Punitorio

    clubes = Club.objects.all().order_by("nombre")
    club_id = request.GET.get("club")

    punitorios = Punitorio.objects.filter(estado="pagado").select_related("club", "concepto", "resuelto_por")
    if club_id:
        punitorios = punitorios.filter(club_id=club_id)
    punitorios = punitorios.order_by("-fecha_pago")

    return render(request, "federacion_app/historial_punitorios.html", {
        "punitorios": punitorios, "clubes": clubes, "club_id": club_id,
    })


@login_required
@user_passes_test(es_federacion)
def resolver_punitorio(request, punitorio_id):
    from .models import Punitorio
    punitorio = get_object_or_404(Punitorio, id=punitorio_id)
    if request.method == "POST":
        punitorio.estado = "pagado"
        punitorio.resuelto_por = request.user
        punitorio.fecha_pago = date.today()
        punitorio.save()
        messages.success(request, f"Punitorio de {punitorio.club} marcado como pagado.")
    return redirect("punitorios_pendientes")


@login_required
@user_passes_test(es_delegado)
def mis_punitorios(request):
    """El delegado ve los punitorios de su club y sube el comprobante de pago."""
    from .models import Punitorio
    punitorios = Punitorio.objects.filter(club=request.user.club).select_related("concepto")
    return render(request, "federacion_app/mis_punitorios.html", {"punitorios": punitorios})


@login_required
@user_passes_test(es_delegado)
def subir_comprobante_punitorio(request, punitorio_id):
    from .models import Punitorio
    punitorio = get_object_or_404(Punitorio, id=punitorio_id, club=request.user.club)
    if request.method == "POST" and request.FILES.get("archivo"):
        punitorio.comprobante_pago = request.FILES["archivo"]
        punitorio.save()
        messages.success(request, "Comprobante subido. La federación lo va a revisar.")
    return redirect("mis_punitorios")
